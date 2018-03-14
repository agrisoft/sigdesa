import datetime
import json
import os
import random
import shutil
import string
import sys
import subprocess
import fileinput
from subprocess import DEVNULL, PIPE, STDOUT, Popen, call
from textwrap import dedent
import config
import base64
import ast
import operator
import zipfile
import ogr2ogr
import proxypy
from dbfread import DBF
from app import app, db
from dictalchemy.utils import arg_to_dict, asdict
from flask import (Response, abort, flash, jsonify, redirect, render_template,
                   request, url_for, session, make_response, send_from_directory)
from flask_login import LoginManager, login_user, login_required, logout_user
from flask_paginate import Pagination, get_page_args
from marshmallow import Schema, fields
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
# from PIL import Image
from sqlalchemy import and_, create_engine, func, or_
from sqlalchemy.sql import text as sa_text
from werkzeug.datastructures import ImmutableMultiDict
from werkzeug.utils import secure_filename
from collections import OrderedDict

# from .forms import LoginForm, KategoriForm, SubKategoriForm, DokumenForm
from .forms import DatadasarForm, LoginForm, VariabelForm
# from .models import t_kategori, t_subkategori, t_dokumen
from .models import (t_datadasar, t_kategori, t_kategoribreak,
                     t_kategoribreakSchema, t_user, t_userSchema,t_variabel, t_config)

from config import APP_ROOT, PG_DB, PG_PASSWORD, PG_USER, PREFIX


# engine open
engine = create_engine(config.SQLALCHEMY_DATABASE_URI,pool_size=30, max_overflow=0)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
 
desclanding = t_config.query.filter_by(config_id='deskripsi').first()
landing_judul = t_config.query.filter_by(config_id='landing_judul').first()
landing_logo = t_config.query.filter_by(config_id='landing_logo').first()
landing_car1 = t_config.query.filter_by(config_id='landing_car1').first()
landing_car2 = t_config.query.filter_by(config_id='landing_car2').first()
landing_car3 = t_config.query.filter_by(config_id='landing_car3').first()

@login_manager.user_loader
def load_user(username):
    # username = t_user.query.filter_by(username=user).first()
    # print('USER:',username)
    return t_user.query.filter_by(username=username).first()

def randomstr_generator(size=6, chars=string.ascii_uppercase + string.digits):
    return ''.join(random.choice(chars) for x in range(size))

@app.route(config.PREFIX + '/login', methods=['GET','POST'])
def login():
    form = LoginForm()
    if request.method == 'GET':
        return render_template('dashboard/login.htm', form=form,title=landing_judul.config_desc, landing_logo=landing_logo.config_desc)
    elif request.method == 'POST':
        if form.validate_on_submit():
            user=t_user.query.filter(or_(t_user.username.like(request.form['username']), t_user.kodepum.like(request.form['username']))).first()
            # user=t_user.query.filter_by(username=request.form['username']).first()
            # print(user.username, user.password)
            if user:
                if user.password == request.form['password']:
                    # login_user(user,remember=True)
                    session['logged_in'] = True
                    session['username'] = user.username
                    isnumber = True
                    try:
                        int(user.username)
                    except:
                        isnumber = False
                    if len(user.username) == 2 and isnumber:
                        session['kelas'] = 'provinsi'
                        session['nama'] = user.provinsi
                    if len(user.username) == 4 and isnumber:
                        session['kelas'] = 'kabupaten'
                        session['nama'] = user.kabupaten
                    if len(user.username) == 10 and isnumber:
                        session['kelas'] = 'desa'
                        session['nama'] = user.desa
                    if not isnumber:
                        session['kelas'] = 'ADMIN'
                        session['nama'] = 'ADMIN'
                    flash('Logged')
                    nextpage = request.args.get('next')
                    print('NEXT:',nextpage)
                    # if not is_safe_url(next):
                    #     return abort(400)
                    return redirect(nextpage or url_for('index'))       
                else:
                    flash('Password Salah!')      
                    return redirect(url_for('login'))        
            else:
                flash('User tidak terdaftar')      
                return redirect(url_for('login'))   
        else:
            flash('Form tidak tervalidasi')
            return redirect(url_for('login'))   

@app.route(config.PREFIX + "/logout")
# @login_required
def logout():
    session['logged_in'] = False
    session['username'] = ''
    session['kelas'] = '' 
    session['nama'] = ''
    # logout_user()
    return redirect(url_for('index'))

@app.route(config.PREFIX + '/')
@app.route(config.PREFIX + '/index')
def index():
    # variabel = t_variabel.query.order_by(t_variabel.id.asc()).all()
    return render_template('dashboard/landing.htm', title=landing_judul.config_desc, landing_logo=landing_logo.config_desc, landing_judul=landing_judul.config_desc, desclanding=desclanding, landing_car1=landing_car1.config_desc,landing_car2=landing_car2.config_desc,landing_car3=landing_car3.config_desc)

@app.route(config.PREFIX + '/profiler')
def profiler():    
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        variabel = t_variabel.query.order_by(t_variabel.id.asc()).all()
        return render_template('dashboard/profiler.htm', title=landing_judul.config_desc, landing_logo=landing_logo.config_desc, landing_judul=landing_judul.config_desc,variabel=variabel)

@app.route(config.PREFIX + '/profiler2')
def profiler2():    
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        variabel = t_variabel.query.order_by(t_variabel.id.asc()).all()
        return render_template('dashboard/profiler2.htm', title=landing_judul.config_desc,landing_logo=landing_logo.config_desc, landing_judul=landing_judul.config_desc,variabel=variabel)

@app.route(config.PREFIX + '/landingedit', methods=['GET', 'POST'])
# @login_required
def landingedit():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        desclanding = t_config.query.filter_by(config_id='deskripsi').first()
        landing_logo = t_config.query.filter_by(config_id='landing_logo').first()
        landing_judul = t_config.query.filter_by(config_id='landing_judul').first()
        landing_car1 = t_config.query.filter_by(config_id='landing_car1').first()
        landing_car2 = t_config.query.filter_by(config_id='landing_car2').first()
        landing_car3 = t_config.query.filter_by(config_id='landing_car3').first()
        if request.method == 'POST':
            print(request.files)
            print(request.form['landing_judul'])
            landingdesc = request.form['deskripsi']
            landingjudul = request.form['landing_judul']
            sqldesc = "UPDATE t_config SET config_desc='%s' WHERE config_id='deskripsi'" % (landingdesc.replace("'","''"))
            engine.execute(sa_text(sqldesc).execution_options(autocommit=True))
            sqljuduldesc = "UPDATE t_config SET config_desc='%s' WHERE config_id='landing_judul'" % (landingjudul.replace("'","''"))
            engine.execute(sa_text(sqljuduldesc).execution_options(autocommit=True))
            if request.files['landing_logo'].filename != '':
                filelogo = request.files['landing_logo']
                encodedfilelogo = base64.b64encode(filelogo.read())
                sqlfilelogo = "UPDATE t_config SET config_desc='%s' WHERE config_id='landing_logo'" % (encodedfilelogo.decode('utf8'))
                engine.execute(sa_text(sqlfilelogo).execution_options(autocommit=True))
            if request.files['landing_car1'].filename != '':
                filecar1 = request.files['landing_car1']
                encodedfilecar1 = base64.b64encode(filecar1.read())
                sqlfilecar1 = "UPDATE t_config SET config_desc='%s' WHERE config_id='landing_car1'" % (encodedfilecar1.decode('utf8'))
                engine.execute(sa_text(sqlfilecar1).execution_options(autocommit=True))
            if request.files['landing_car2'].filename != '':
                filecar2 = request.files['landing_car2']
                encodedfilecar2 = base64.b64encode(filecar2.read())
                sqlfilecar2 = "UPDATE t_config SET config_desc='%s' WHERE config_id='landing_car2'" % (encodedfilecar2.decode('utf8'))
                engine.execute(sa_text(sqlfilecar2).execution_options(autocommit=True))
            if request.files['landing_car3'].filename != '':
                filecar3 = request.files['landing_car3']
                encodedfilecar3 = base64.b64encode(filecar3.read())
                sqlfilecar3 = "UPDATE t_config SET config_desc='%s' WHERE config_id='landing_car3'" % (encodedfilecar3.decode('utf8'))
                engine.execute(sa_text(sqlfilecar3).execution_options(autocommit=True))
            return redirect(url_for('index'))
        return render_template('dashboard/landingedit.htm', title=landing_judul.config_desc, landing_logo=landing_logo.config_desc,desclanding=desclanding, landing_judul=landing_judul)

@app.route(config.PREFIX + '/peta')
def peta():
    return render_template('dashboard/peta.htm', landing_logo=landing_logo.config_desc, landing_judul=landing_judul.config_desc,title=landing_judul.config_desc,)

@app.route(config.PREFIX + '/profil_provinsi')
def profil_provinsi():
    variabel = t_variabel.query.order_by(t_variabel.id.asc()).all()
    return render_template('dashboard/profil_provinsi.htm', title=landing_judul.config_desc,landing_logo=landing_logo.config_desc, landing_judul=landing_judul.config_desc, variabel=variabel)

@app.route(config.PREFIX + '/variabel')
def variabel():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        offset = get_page_args()
        cari = request.args.get('cari')
        cariteks = request.args.get('cariTeks')
        if cari:
            string_cari = '%' + cari + '%'
            rows = t_variabel.query.filter(or_(t_variabel.n_variabel.ilike(string_cari), t_variabel.r_variabel.ilike(string_cari))).count()
            variabel = t_variabel.query.filter(or_(t_variabel.n_variabel.ilike(string_cari), t_variabel.r_variabel.ilike(string_cari))).order_by(t_variabel.r_variabel.asc()).offset(offset[2]).limit(offset[1])
        else:
            rows = t_variabel.query.order_by(t_variabel.n_variabel).count()
            variabel = t_variabel.query.order_by(t_variabel.r_variabel.asc()).offset(offset[2]).limit(offset[1])
        per_page = 10
        search = False
        q = request.args.get('q')
        if q:
            search = True
        page = request.args.get('page', type=int, default=1)
        pagination = Pagination(page=page, total=rows, per_page=per_page, search=search, record_name='Variabel',css_framework='bootstrap3')
        return render_template('dashboard/variabel.htm',title=landing_judul.config_desc, landing_logo=landing_logo.config_desc, landing_judul=landing_judul.config_desc,variabel=variabel, per_page=per_page,  page=page, pagination=pagination)

@app.route(config.PREFIX + '/variabelbaru', methods=['GET', 'POST'])
def variabelbaru():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        form = VariabelForm()
        if form.validate_on_submit():
            if request.method == 'POST':
                variabel = t_variabel(n_variabel=request.form['n_variabel'])
                variabel.r_variabel = request.form['r_variabel']
                variabel.deskripsi = request.form['deskripsi']
                # r_variabel = t_variabel(r_variabel=request.form['r_variabel'])
                db.session.add(variabel)
                db.session.commit()
                # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
                sqlcreatetd = 'CREATE TABLE IF NOT EXISTS public.td_%s (id BIGSERIAL PRIMARY KEY, indikator character varying(64), value text)' % (request.form['n_variabel'])
                print(sqlcreatetd)
                engine.execute(sqlcreatetd)
                return redirect(url_for('variabel'))
        return render_template('dashboard/variabelbaru.htm', landing_logo=landing_logo.config_desc, landing_judul=landing_judul.config_desc,form=form,title=landing_judul.config_desc)

@app.route(config.PREFIX + '/variabel/delete/<n_variabel>', methods=['GET', 'POST'])
def variabeldelete(n_variabel):
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        variabel = t_variabel.query.filter_by(n_variabel=n_variabel).first()
        if request.method == 'POST':
            # t_variabel.query.filter_by(n_variabel=n_variabel).delete()
            db.session.delete(variabel)
            db.session.commit()
            sqldelview = "DROP VIEW IF EXISTS tv_%s" % (n_variabel)
            sqldelgview = "DROP VIEW IF EXISTS tvg_%s" % (n_variabel)
            sqldeldesc = "DROP TABLE IF EXISTS td_%s" % (n_variabel)
            sqldelisian = "DROP TABLE IF EXISTS ti_%s" % (n_variabel)
            sqldelskor = "DROP TABLE IF EXISTS ts_%s" % (n_variabel)
            sqldelintv =  "DROP TABLE IF EXISTS tiv_%s" % (n_variabel)
            engine.execute(sa_text(sqldelview).execution_options(autocommit=True))
            engine.execute(sa_text(sqldelgview).execution_options(autocommit=True))
            engine.execute(sa_text(sqldeldesc).execution_options(autocommit=True))
            engine.execute(sa_text(sqldelisian).execution_options(autocommit=True))
            engine.execute(sa_text(sqldelskor).execution_options(autocommit=True))
            engine.execute(sa_text(sqldelintv).execution_options(autocommit=True))
            return redirect(url_for('variabel'))
        return render_template('dashboard/variabeldelete.htm', landing_logo=landing_logo.config_desc, landing_judul=landing_judul.config_desc,variabel=variabel,title=landing_judul.config_desc)

@app.route(config.PREFIX + '/variabel/edit/<n_variabel>', methods=['GET', 'POST'])
def variabeledit(n_variabel):
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        print(request.form)
        form = VariabelForm()
        variabel = t_variabel.query.filter_by(n_variabel=n_variabel).first()
        if request.method == 'POST':
            variabel.r_variabel = request.form['r_variabel']
            variabel.deskripsi = request.form['deskripsi']
            db.session.commit()
            return redirect(url_for('variabel'))
        else:
            return render_template('dashboard/variabeledit.htm', landing_logo=landing_logo.config_desc, landing_judul=landing_judul.config_desc,variabel=variabel, form=form,title=landing_judul.config_desc)

@app.route(config.PREFIX + '/variabel/delete/')
def variabeldelete_dum():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        return render_template('dashboard/variabeldelete.htm',landing_logo=landing_logo.config_desc, landing_judul=landing_judul.config_desc,title=landing_judul.config_desc)   

@app.route(config.PREFIX + '/variabel/edit/')
def variabeledit_dum():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        return render_template('dashboard/variabeledit.htm',landing_logo=landing_logo.config_desc, landing_judul=landing_judul.config_desc,title=landing_judul.config_desc)   

@app.route(config.PREFIX + '/variabel/save/<id>', methods=['GET', 'POST'])
def variabelsave():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        if request.method == 'POST':
            pass
        return redirect(url_for(variabel) + '/' + str(id))

@app.route(config.PREFIX + '/variabel/<id>')
def variabeldetail(id):
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        variabel = t_variabel.query.filter_by(id=id).first()
        return render_template('dashboard/variabeldetail.htm', landing_logo=landing_logo.config_desc, landing_judul=landing_judul.config_desc,variabel=variabel,title=landing_judul.config_desc)

@app.route(config.PREFIX + '/variabel/detail/<id>', methods=['GET'])
def variabeldetailload(id):
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        table = 'td_' + str(id)
        output = {}
        # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
        result = engine.execute('select * from ' + table + ' order by id')
        for item in result:
            print(item.indikator, item.value)
            output[item.indikator] = item.value
        jsonoutput = json.dumps(output, sort_keys=True)
        print(jsonoutput)
        # engine.dispose()
        return Response(jsonoutput, mimetype='application/json')

@app.route(config.PREFIX + '/datadasar')
def datadasar():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        offset = get_page_args()
        cari = request.args.get('cari')
        cariteks = request.args.get('cariTeks')
        if cari:
            string_cari = '%' + cari + '%'
            rows = t_datadasar.query.filter(or_(t_datadasar.n_datadasar.ilike(string_cari), t_datadasar.r_datadasar.ilike(string_cari))).count()
            datadasar = t_datadasar.query.filter(or_(t_datadasar.n_datadasar.ilike(string_cari), t_datadasar.r_datadasar.ilike(string_cari))).order_by(t_datadasar.r_datadasar.asc()).offset(offset[2]).limit(offset[1])
        else:
            rows = t_datadasar.query.order_by(t_datadasar.n_datadasar).count()
            datadasar = t_datadasar.query.order_by(t_datadasar.r_datadasar.asc()).offset(offset[2]).limit(offset[1])
            # for item in datadasar:
            #     print(item)
        per_page = 10
        search = False
        q = request.args.get('q')
        if q:
            search = True
        page = request.args.get('page', type=int, default=1)
        pagination = Pagination(page=page, total=rows, per_page=per_page, search=search, record_name='Data Dasar',css_framework='bootstrap3')
        return render_template('dashboard/datadasar.htm', landing_logo=landing_logo.config_desc, landing_judul=landing_judul.config_desc,title=landing_judul.config_desc, datadasar=datadasar, per_page=per_page,  page=page, pagination=pagination)

@app.route(config.PREFIX +'/datadasarget', methods=['GET'])
def datadasarget():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        table = 't_datadasar'
        output = {}
        # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
        result = engine.execute('select * from ' + table)
        for item in result:
            output[item.n_datadasar] = item.r_datadasar
        jsonoutput = json.dumps(output)
        # engine.dispose()
        return Response(jsonoutput, mimetype='application/json')

@app.route(config.PREFIX +'/datadasardetail')
def datadasardetail():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:      
        offset = get_page_args()
        cari = request.args.get('cari')
        cariteks = request.args.get('cariTeks')
        tabel = request.args['n_datadasar']
        # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
        sql1 = "SELECT * FROM tb_%s OFFSET %s LIMIT %s" % (tabel,offset[2],offset[1])
        sql2 = "SELECT * FROM tb_%s" % (tabel)
        res = engine.execute(sa_text(sql1))
        resn = engine.execute(sa_text(sql2))
        result = res.fetchall()
        result2 = resn.fetchall()
        nresult = len(result2)
        res2 = engine.execute("select column_name,data_type from INFORMATION_SCHEMA.COLUMNS where table_name='tb_" + tabel + "'")
        # result2 = res.fetchall()
        # print('select * from tb_' + tabel)
        # for item in result:
        #     print(item)
        # for item in result2:
        #     print(item)
        # engine.dispose()        
        per_page = 10
        search = False
        q = request.args.get('q')
        if q:
            search = True
        page = request.args.get('page', type=int, default=1)
        pagination = Pagination(page=page, total=nresult, per_page=per_page, search=search, record_name='Data Dasar',css_framework='bootstrap3')
        return render_template('dashboard/datadasardetail.htm', landing_logo=landing_logo.config_desc, landing_judul=landing_judul.config_desc,tabel=tabel,result=result, result2=res2, per_page=per_page,  page=page, pagination=pagination,title=landing_judul.config_desc)

@app.route(config.PREFIX +'/datadasardesc', methods=['GET'])
def datadasardesc():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        table = 'tb_' + request.args.get('iddata')
        output = {}
        # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
        result = engine.execute("select column_name,data_type from INFORMATION_SCHEMA.COLUMNS where table_name='" + table + "'")
        print(result)
        for item in result:
            output[item.column_name] = item.data_type
        jsonoutput = json.dumps(output)
        # engine.dispose()
        return Response(jsonoutput, mimetype='application/json')    


@app.route(config.PREFIX +'/datadasarvdesc', methods=['GET'])
def datadasarvdesc():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        tabel = request.args['tabel']
        # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
        sqlgetdata = "SELECT * from tbd_%s order by variabel" % (tabel)
        print(sqlgetdata)
        result = engine.execute(sa_text(sqlgetdata)).fetchall()
        output = {}
        try:
            for val in result:
                output[val[1]] = val[2]
        except:
            pass
        jsonoutput = json.dumps(output)
        # engine.dispose()
        return Response(jsonoutput, mimetype='application/json')

@app.route(config.PREFIX +'/dasarmapcoloragg', methods=['GET'])
def dasarmapcoloragg():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        tabel = request.args['tabel']
        tahun = request.args['tahun']
        kolom = request.args['kolom']
        # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
        sqlget = "SELECT DISTINCT SUBSTRING(kdebps,1,4), SUM(%s) FROM tb_%s WHERE tahun=%s GROUP BY SUBSTRING(kdebps,1,4)" % (kolom,tabel,tahun)
        # sqlgetgraph = "SELECT kategori, SUM(count) as jumlah from tv_%s WHERE tahun=%s GROUP BY kategori ORDER BY kategori" % (tabel,tahun)
        print(sqlget)
        result = engine.execute(sa_text(sqlget))
        output = []
        for item in result:
            inner = {}
            jml = str(item[1])
            inner['kdbbps'] = item[0]
            inner['isi'] = jml
            output.append(inner)
        # engine.dispose()
        jsonoutput = json.dumps(output)
        return Response(jsonoutput, mimetype='application/json')

@app.route(config.PREFIX +'/dasarmapcolor', methods=['GET'])
def dasarmapcolor():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        tabel = request.args['tabel']
        tahun = request.args['tahun']
        kolom = request.args['kolom']
        kode = request.args['kode']
        # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
        sqlget = "SELECT DISTINCT kdebps, SUM(%s) FROM tb_%s WHERE tahun=%s AND SUBSTRING(kdebps,1,4)='%s' GROUP BY kdebps" % (kolom,tabel,tahun,kode)
        # sqlgetgraph = "SELECT kategori, SUM(count) as jumlah from tv_%s WHERE tahun=%s GROUP BY kategori ORDER BY kategori" % (tabel,tahun)
        print(sqlget)
        result = engine.execute(sa_text(sqlget))
        output = []
        for item in result:
            inner = {}
            jml = str(item[1])
            inner['kdebps'] = item[0]
            inner['isi'] = jml
            output.append(inner)
        # engine.dispose()
        jsonoutput = json.dumps(output)
        return Response(jsonoutput, mimetype='application/json')

@app.route(config.PREFIX +'/datadasarbaru', methods=['POST','GET'])
def datadasarbaru():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        form = DatadasarForm()
        if form.validate_on_submit():
            if request.method == 'POST':
                datadasar = t_datadasar(n_datadasar=request.form['n_datadasar'])
                datadasar.r_datadasar = request.form['r_datadasar']
                # db.session.add(datadasar)
                # db.session.commit()
                return redirect(url_for('datadasar'))
        return render_template('dashboard/datadasarbaru.htm', landing_logo=landing_logo.config_desc, landing_judul=landing_judul.config_desc,form=form,title=landing_judul.config_desc)

@app.route(config.PREFIX +'/datadasarhandler', methods=['POST'])
def datadasarhandler():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        print(dict(request.form))
        if request.method == 'POST':
            # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
            data = dict(request.form)
            newtable = data['n_datadasar'][0]
            kettable = data['r_datadasar'][0]
            kolnum = len(data['namakolom'])
            # base_sql = "CREATE TABLE IF NOT EXISTS public.tb_%s (id BIGSERIAL PRIMARY KEY, kdebps character varying(12), kdepum character varying(12), tahun integer)" % (newtable)
            base_sql = "CREATE TABLE IF NOT EXISTS public.tb_%s (id BIGSERIAL PRIMARY KEY, kdebps character varying(12), tahun integer)" % (newtable)
            engine.execute(base_sql)
            base_sql2 = "CREATE TABLE IF NOT EXISTS public.tbd_%s (id BIGSERIAL PRIMARY KEY, variabel text, v_deskripsi text)" % (newtable)
            engine.execute(base_sql2)
            print(base_sql)
            for i in range(kolnum):
                if data['tipe'][i] == '1':
                    tipedata = 'text'
                if data['tipe'][i] == '2':
                    tipedata = 'real'
                col_sql = "ALTER TABLE public.tb_%s ADD COLUMN %s %s" % (newtable,data['namakolom'][i],tipedata)
                print(col_sql)
                engine.execute(col_sql)
                if tipedata == 'real':
                    row_sql = "INSERT INTO public.tbd_%s (variabel, v_deskripsi) values ('%s','%s')" % (newtable,data['namakolom'][i],data['keterangan'][i])
                    engine.execute(row_sql)
            t_sql = "INSERT INTO t_datadasar (n_datadasar, r_datadasar) values ('%s', '%s')" % (newtable, kettable)
            engine.execute(t_sql)
            # engine.dispose()
            return Response("{success:true}", mimetype='application/json')

@app.route(config.PREFIX +'/datadasar/edit', methods=['POST','GET'])
def datadasaredit():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        form = DatadasarForm()
        n_datadasar = request.args.get('n_datadasar')
        datadasar = t_datadasar.query.filter_by(n_datadasar=n_datadasar).first()
        # if form.validate_on_submit():
        if request.method == 'POST':
            datadasar.r_datadasar = request.form['r_datadasar']
            print("args",datadasar.r_datadasar)
            db.session.commit()
            return redirect(url_for('datadasar'))
        return render_template('dashboard/datadasaredit.htm', landing_logo=landing_logo.config_desc, landing_judul=landing_judul.config_desc,form=form, datadasar=datadasar,title=landing_judul.config_desc)

@app.route(config.PREFIX +'/datadasar/delete', methods=['POST'])
def datadasardelete():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        pass

@app.route(config.PREFIX +'/dddelete', methods=['GET','POST'])
def dddelete():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        if request.method == 'POST':
            tabel = request.args.get('tabel')
            sqldel1 = "DELETE FROM t_datadasar WHERE n_datadasar='%s'" % (tabel)
            engine.execute(sa_text(sqldel1).execution_options(autocommit=True))
            sqldel2 = "DROP TABLE IF EXISTS tb_%s" % (tabel)
            engine.execute(sa_text(sqldel2).execution_options(autocommit=True))
            sqldel3 = "DROP TABLE IF EXISTS tbd_%s" % (tabel)
            engine.execute(sa_text(sqldel3).execution_options(autocommit=True))
            return redirect(url_for('datadasar'))
        else:
            n_datadasar = request.args.get('tabel')
            datadasar = t_datadasar.query.filter_by(n_datadasar=n_datadasar).first()
            return render_template('dashboard/datadasardelete.htm', landing_logo=landing_logo.config_desc, landing_judul=landing_judul.config_desc, datadasar=datadasar,title=landing_judul.config_desc)

@app.route(config.PREFIX +'/daftar_desa')
def daftar_desa():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        offset = get_page_args()
        cari = request.args.get('cari')
        cariteks = request.args.get('cariTeks')
        loggeduser = session.get('username')
        cariuser = '%' + loggeduser + '%'
        print(loggeduser)
        if cari:
            string_cari = '%' + cari + '%'
            rows = t_user.query.filter(or_(t_user.username.ilike(string_cari), t_user.kodepum.ilike(string_cari), t_user.desa.ilike(string_cari), t_user.kecamatan.ilike(string_cari), t_user.kabupaten.ilike(string_cari))).count()
            desa = t_user.query.filter(or_(t_user.username.ilike(string_cari), t_user.kodepum.ilike(string_cari), t_user.desa.ilike(string_cari), t_user.kecamatan.ilike(string_cari), t_user.kabupaten.ilike(string_cari))).order_by(t_user.username.asc()).offset(offset[2]).limit(offset[1])
        else:
            if loggeduser == 'administrator':
                rows = t_user.query.count()
                desa = t_user.query.offset(offset[2]).limit(offset[1])
            else:
                rows = t_user.query.filter(t_user.username.ilike(cariuser)).count()
                desa = t_user.query.filter(t_user.username.ilike(cariuser)).offset(offset[2]).limit(offset[1])
        per_page = 10
        search = False
        q = request.args.get('q')
        if q:
            search = True
        page = request.args.get('page', type=int, default=1)
        pagination = Pagination(page=page, total=rows, per_page=per_page, search=search, record_name='Desa',css_framework='bootstrap3')
        return render_template('dashboard/list_desa.htm', landing_logo=landing_logo.config_desc, landing_judul=landing_judul.config_desc,title=landing_judul.config_desc, desa=desa, per_page=per_page,  page=page, pagination=pagination)

@app.route(config.PREFIX +'/isivariabel')
def isivariabel():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        variabel = t_variabel.query.order_by(t_variabel.id.asc()).all()
        # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
        provinsi = "select distinct provinsi, left(username, 2) as username from t_user where provinsi != 'ADMIN' and left(username, 2) !='' and username !='0' order by left(username, 2)" 
        kabupaten = "select distinct kabupaten, left(username, 4) as username from t_user where provinsi != 'ADMIN' and left(username, 4) !='' and kabupaten !='-' and username !='0' order by kabupaten"
        desa = "select distinct desa, username from t_user where provinsi != 'ADMIN' order by desa"
        l_provinsi = engine.execute(provinsi)
        l_kabupaten = engine.execute(kabupaten)
        l_desa = engine.execute(desa)
        # engine.dispose()
        return render_template('dashboard/isivariabel.htm', landing_logo=landing_logo.config_desc, landing_judul=landing_judul.config_desc,variabel=variabel, provinsi=l_provinsi, kabupaten=l_kabupaten, desa=l_desa,title=landing_judul.config_desc)

@app.route(config.PREFIX +'/simpanvariabel', methods=['POST'])
def simpanvariabel():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        if request.method == 'POST':
            tabel = request.args['tabel']
            tipe = request.args['tipe']
            data = dict(request.form)
            print('SAVE VAR: ',tabel, data)
            # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
            param = ''
            keterangan = ''
            sqldroptable = "DROP TABLE IF EXISTS public.ti_%s" % (tabel)
            engine.execute(sqldroptable)
            sqlcreatetabel = "CREATE TABLE IF NOT EXISTS public.ti_%s (id BIGSERIAL PRIMARY KEY, kdebps character varying(12), kdepum character varying(12), tahun smallint, skor real)" % (tabel)
            engine.execute(sa_text(sqlcreatetabel).execution_options(autocommit=True))
            # engine.execute(sqlcreatetabel)
            jumlah = 0
            for item in data:
                jumlah = jumlah + 1
                sqladdcol = "ALTER TABLE public.ti_%s ADD COLUMN %s %s" % (tabel,item,'boolean')
                engine.execute(sqladdcol)
            # sqlskor = "ALTER TABLE public.ti_%s ADD COLUMN skor integer" % (tabel)
            # engine.execute(sqlskor)
            sqltruncate = "TRUNCATE TABLE td_%s CASCADE" % (tabel)
            print(sqltruncate)
            engine.execute(sa_text(sqltruncate).execution_options(autocommit=True))
            sqlinsert = "INSERT INTO td_%s (indikator, value) values ('tipe', '%s')" % (tabel, str(tipe))
            print(sqlinsert)
            engine.execute(sqlinsert)
            for item in data:
                print(item, data[item][0])
                sqlinsert = "INSERT INTO td_%s (indikator, value) values ('%s', '%s')" % (tabel, item, data[item][0])
                print(sqlinsert)
                engine.execute(sqlinsert)
            # try:
            #     sqlclearkategori = "DELETE FROM t_kategoribreak WHERE n_variabel='%s'" % (tabel)
            #     engine.execute(sqlclearkategori)
            # except:
            #     pass
            sqlkategori = "INSERT INTO t_kategoribreak (n_variabel, n_break, v_break, v_color) values ('%s',%s,'%s','%s')" % (tabel, 1, jumlah, '#123123')
            engine.execute(sqlkategori)
            sqltskort = "CREATE TABLE IF NOT EXISTS public.ts_%s (id BIGSERIAL PRIMARY KEY, kdebps character varying(12), kdepum character varying(12), tahun smallint, skor real, kategori text)" % (tabel) 
            engine.execute(sa_text(sqltskort).execution_options(autocommit=True))
            sqlagg = "CREATE OR REPLACE VIEW tv_%s AS SELECT DISTINCT a.kdbbps, a.count, a.kategori, b.tahun FROM (SELECT z.kdbbps, z.count, y.kategori FROM (SELECT t.kdbbps, MAX(t.count) AS count FROM (SELECT DISTINCT kategori,SUBSTRING(kdebps,1,4) AS kdbbps, COUNT(SUBSTRING(kdebps,1,4)) AS count, tahun FROM ts_%s GROUP BY kdbbps, tahun, kategori ORDER BY kdbbps) AS t GROUP BY t.kdbbps ORDER BY t.kdbbps) AS z, (SELECT t.kdbbps, t.kategori, MAX(t.count) AS count FROM (SELECT DISTINCT kategori,SUBSTRING(kdebps,1,4) AS kdbbps, COUNT(SUBSTRING(kdebps,1,4)) AS count, tahun FROM ts_%s GROUP BY kdbbps, tahun, kategori ORDER BY kdbbps) AS t GROUP BY t.kdbbps, t.kategori ORDER BY t.kdbbps)  AS y WHERE z.count = y.count AND y.kdbbps = z.kdbbps ORDER BY z.kdbbps) AS a, ts_%s AS b WHERE a.kdbbps = SUBSTRING(b.kdebps,1,4) ORDER BY a.kdbbps" % (tabel, tabel, tabel, tabel)
            engine.execute(sqlagg)
            sqlcreategview = "CREATE OR REPLACE VIEW public.tvg_%s AS SELECT a.geom, b.* FROM public.admin_desa a FULL OUTER JOIN public.ts_%s b ON a.kdebps=b.kdebps" % (tabel,tabel)
            engine.execute(sqlcreategview)
            # engine.dispose()
            # return Response("{OK}", mimetype='application/json') 
            return redirect(url_for('variabel'))

@app.route(config.PREFIX +'/simpanvariabel2', methods=['POST'])
def simpanvariabel2():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        if request.method == 'POST':
            tabel = request.args['tabel']
            tipe = request.args['tipe']
            pil = request.args['pil']
            data = dict(request.form)
            print('SAVE VAR: ',tabel, data)
            param = ''
            keterangan = ''
            sqldroptable = "DROP TABLE IF EXISTS public.ti_%s" % (tabel)
            engine.execute(sqldroptable)
            sqlcreatetabel = "CREATE TABLE IF NOT EXISTS public.ti_%s (id BIGSERIAL PRIMARY KEY, kdebps character varying(12), kdepum character varying(12), tahun smallint, skor real)" % (tabel)
            engine.execute(sa_text(sqlcreatetabel).execution_options(autocommit=True))
            # engine.execute(sqlcreatetabel)
            parent = []
            child = []
            subchild = {}
            jmlchild = 0
            for item in data:
                if item.split('_')[0] == 'pilgan':
                    tmp = {}
                    tmp[1] = data[item][0]
                    subchild[int(item.split('_')[1])] = tmp
                    jmlchild = jmlchild + 1
                    child.append(subchild)
                    sqladdcol = "ALTER TABLE public.ti_%s ADD COLUMN %s %s" % (tabel,item,'real')
                    engine.execute(sqladdcol)
            print("SUB",subchild)
            for item in subchild:
                print(item)
                for subitem in data:
                    # print("I",subitem)
                    u = subitem.split('_')
                    try:
                        a = u[0]
                        b = int(u[1])
                        c = int(u[2])
                        print("J",u[0],u[1],u[2],data[subitem][0])
                        if item == b:
                            subchild[b][c+1] = data[subitem][0]
                        # else:
                        #     subchild[b][c] = data[subitem][0]
                    except:
                        pass
                        # print("L",subitem)
            print("SUB",subchild)
            # sqlskor = "ALTER TABLE public.ti_%s ADD COLUMN skor integer" % (tabel)
            # engine.execute(sqlskor)
            sqltruncate = "TRUNCATE TABLE td_%s CASCADE" % (tabel)
            print(sqltruncate)
            engine.execute(sa_text(sqltruncate).execution_options(autocommit=True))
            sqlinsert = "INSERT INTO td_%s (indikator, value) values ('tipe', '%s')" % (tabel, str(tipe))
            print(sqlinsert)
            engine.execute(sqlinsert)
            try:
                sqldel = "DELETE FROM td_%s WHERE indikator='pertanyaan'" % (tabel)
                engine.execute(sa_text(sqldel))
                sql = "INSERT INTO td_%s (indikator, value) values ('pertanyaan','%s')" % (tabel, json.dumps(subchild))
                engine.execute(sa_text(sql))
            except:
                sql = "INSERT INTO td_%s (indikator, value) values ('pertanyaan','%s')" % (tabel, json.dumps(subchild))
                engine.execute(sa_text(sql))
            # try:
            #     sqlclearkategori = "DELETE FROM t_kategoribreak WHERE n_variabel='%s'" % (tabel)
            #     engine.execute(sqlclearkategori)
            # except:
            #     pass
            sqlkategori = "INSERT INTO t_kategoribreak (n_variabel, n_break, v_break, v_color) values ('%s',%s,'%s','%s')" % (tabel, 1, 1, '#123123')
            engine.execute(sqlkategori)
            sqltskort = "CREATE TABLE IF NOT EXISTS public.ts_%s (id BIGSERIAL PRIMARY KEY, kdebps character varying(12), kdepum character varying(12), tahun smallint, skor real, kategori text)" % (tabel) 
            engine.execute(sa_text(sqltskort).execution_options(autocommit=True))
            sqlagg = "CREATE OR REPLACE VIEW tv_%s AS SELECT DISTINCT a.kdbbps, a.count, a.kategori, b.tahun FROM (SELECT z.kdbbps, z.count, y.kategori FROM (SELECT t.kdbbps, MAX(t.count) AS count FROM (SELECT DISTINCT kategori,SUBSTRING(kdebps,1,4) AS kdbbps, COUNT(SUBSTRING(kdebps,1,4)) AS count, tahun FROM ts_%s GROUP BY kdbbps, tahun, kategori ORDER BY kdbbps) AS t GROUP BY t.kdbbps ORDER BY t.kdbbps) AS z, (SELECT t.kdbbps, t.kategori, MAX(t.count) AS count FROM (SELECT DISTINCT kategori,SUBSTRING(kdebps,1,4) AS kdbbps, COUNT(SUBSTRING(kdebps,1,4)) AS count, tahun FROM ts_%s GROUP BY kdbbps, tahun, kategori ORDER BY kdbbps) AS t GROUP BY t.kdbbps, t.kategori ORDER BY t.kdbbps)  AS y WHERE z.count = y.count AND y.kdbbps = z.kdbbps ORDER BY z.kdbbps) AS a, ts_%s AS b WHERE a.kdbbps = SUBSTRING(b.kdebps,1,4) ORDER BY a.kdbbps" % (tabel, tabel, tabel, tabel)
            engine.execute(sqlagg)
            sqlcreategview = "CREATE OR REPLACE VIEW public.tvg_%s AS SELECT a.geom, b.* FROM public.admin_desa a FULL OUTER JOIN public.ts_%s b ON a.kdebps=b.kdebps" % (tabel,tabel)
            engine.execute(sqlcreategview)
            # return Response("{OK}", mimetype='application/json') 
            return redirect(url_for('variabel'))

@app.route(config.PREFIX +'/simpanvariabel3', methods=['POST'])
def simpanvariabel3():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        if request.method == 'POST':
            tabel = request.args['tabel']
            tipe = request.args['tipe']
            # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
            param = ''
            keterangan = ''
            sqldroptable = "DROP TABLE IF EXISTS public.ti_%s" % (tabel)
            engine.execute(sqldroptable)
            sqlcreatetabel = "CREATE TABLE IF NOT EXISTS public.ti_%s (id BIGSERIAL PRIMARY KEY, kdebps character varying(12), kdepum character varying(12), tahun smallint, skor real)" % (tabel)
            engine.execute(sa_text(sqlcreatetabel).execution_options(autocommit=True))
            # engine.execute(sqlcreatetabel)
            # jumlah = 0
            # for item in data:
            #     jumlah = jumlah + 1
                # sqladdcol = "ALTER TABLE public.ti_%s ADD COLUMN %s %s" % (tabel,item,'boolean')
            #     engine.execute(sqladdcol)
            sqlskor = "ALTER TABLE public.ti_%s ADD COLUMN htmldata text" % (tabel)
            engine.execute(sqlskor)
            sqltruncate = "TRUNCATE TABLE td_%s CASCADE" % (tabel)
            print(sqltruncate)
            engine.execute(sa_text(sqltruncate).execution_options(autocommit=True))
            sqlinsert = "INSERT INTO td_%s (indikator, value) values ('tipe', '%s')" % (tabel, str(tipe))
            print(sqlinsert)
            engine.execute(sqlinsert)
            # try:
            #     sqlclearkategori = "DELETE FROM t_kategoribreak WHERE n_variabel='%s'" % (tabel)
            #     engine.execute(sqlclearkategori)
            # except:
            #     pass
            sqlkategori = "INSERT INTO t_kategoribreak (n_variabel, n_break, v_break, v_color) values ('%s',%s,'%s','%s')" % (tabel, 1, 1, '#123123')
            engine.execute(sqlkategori)
            sqltskort = "CREATE TABLE IF NOT EXISTS public.ts_%s (id BIGSERIAL PRIMARY KEY, kdebps character varying(12), kdepum character varying(12), tahun smallint, skor real, kategori text)" % (tabel) 
            engine.execute(sa_text(sqltskort).execution_options(autocommit=True))
            sqlagg = "CREATE OR REPLACE VIEW public.tv_%s AS SELECT DISTINCT SUBSTRING(kdebps,1,4) AS kdbbps, COUNT(SUBSTRING(kdebps,1,4)) AS count, tahun FROM ts_%s GROUP BY kdbbps, tahun" % (tabel, tabel)
            engine.execute(sqlagg)
            sqlcreategview = "CREATE OR REPLACE VIEW public.tvg_%s AS SELECT a.geom, b.* FROM public.admin_desa a FULL OUTER JOIN public.ts_%s b ON a.kdebps=b.kdebps" % (tabel,tabel)
            engine.execute(sqlcreategview)
            # for item in data:
            #     print(item, data[item][0])
            #     sqlinsert = "INSERT INTO td_%s (indikator, value) values ('%s', '%s')" % (tabel, item, data[item][0])
            #     print(sqlinsert)
            #     engine.execute(sqlinsert)
            # sqlkategori = "INSERT INTO t_kategoribreak (n_variabel, n_break, v_break, v_color) values ('%s',%s,'%s','%s')" % (tabel, 1, jumlah, '#123123')
            # engine.execute(sqlkategori)
            # engine.dispose()        
            # return Response("{OK}", mimetype='application/json') 
            return redirect(url_for('variabel'))

@app.route(config.PREFIX +'/simpanvariabel4', methods=['POST'])
def simpanvariabel4():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        if request.method == 'POST':
            tabel = request.args['tabel']
            tipe = request.args['tipe']
            data = dict(request.form)
            print(data)
            formula = data['formula'][0]
            tahun = data['tahun'][0]
            sqldroptable = "DROP TABLE IF EXISTS public.ti_%s" % (tabel)
            engine.execute(sa_text(sqldroptable).execution_options(autocommit=True))
            # sqlcreatetabel = "CREATE TABLE IF NOT EXISTS public.ti_%s (id BIGSERIAL PRIMARY KEY, kdebps character varying(12), kdepum character varying(12), tahun smallint, skor real)" % (tabel)
            # engine.execute(sa_text(sqlcreatetabel).execution_options(autocommit=True))
            sel1 = "SELECT admin_desa.kdebps, (%s) as skor " % (formula)
            sel2 = "INTO ti_%s FROM admin_desa " % (tabel)
            sel3 = " "
            for item in data['tabels[]']:
                # sel2 = sel2 + item + ','
                sel3 = sel3 + "INNER JOIN %s ON admin_desa.kdebps=%s.kdebps " % (item, item)
            sel2 = sel2[:-1]
            sql = sel1 + sel2 + sel3
            print("SQL", sql)
            engine.execute(sa_text(sql).execution_options(autocommit=True))
            # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
            param = ''
            keterangan = ''
            sqltruncate = "TRUNCATE TABLE td_%s CASCADE" % (tabel)
            print(sqltruncate)
            engine.execute(sa_text(sqltruncate).execution_options(autocommit=True))
            sqlinsert = "INSERT INTO td_%s (indikator, value) values ('tipe', '%s')" % (tabel, str(tipe))
            print(sqlinsert)
            engine.execute(sqlinsert)
            try:
                sqldel = "DELETE FROM td_%s WHERE indikator='formula'" % (tabel)
                engine.execute(sa_text(sqldel))
                sqldel2 = "DELETE FROM td_%s WHERE indikator='tabel'" % (tabel)
                engine.execute(sa_text(sqldel2))
                sqla = "INSERT INTO td_%s (indikator, value) values ('formula','%s')" % (tabel, formula)
                engine.execute(sa_text(sqla))
                sqlc = "INSERT INTO td_%s (indikator, value) values ('tabel','%s')" %  (tabel, json.dumps(data['tabels[]']))
                engine.execute(sa_text(sqlc))
                print("A",sqla)
            except:
                sqlb = "INSERT INTO td_%s (indikator, value) values ('formula','%s')" % (tabel, formula)
                engine.execute(sa_text(sqlb))
                sqlc = "INSERT INTO td_%s (indikator, value) values ('tabel','%s')" % (tabel, json.dumps(data['tabels[]']))
                engine.execute(sa_text(sqlc))
                print("B",sqla)
            sqltskort = "CREATE TABLE IF NOT EXISTS public.ts_%s (id BIGSERIAL PRIMARY KEY, kdebps character varying(12), kdepum character varying(12), tahun smallint, skor real, kategori text)" % (tabel) 
            engine.execute(sa_text(sqltskort).execution_options(autocommit=True))
            kategoribreak = t_kategoribreak.query.filter_by(n_variabel=tabel).first()
            kategoribreak_schema = t_kategoribreakSchema()
            output = kategoribreak_schema.dump(kategoribreak).data
            # print(output)
            sqlfor = "SELECT * from ti_%s" % (tabel)
            forresultsql = engine.execute(sqlfor)
            forresult = forresultsql.fetchall()
            for item in forresult:
                # print(item[0],item[1])
                skor = item[1]
                kode = item[0]
                skoring = output['v_break'].split(',')
                r_skoring = output['r_break'].split(',')
                ranges = range(0)
                it = 0
                ra = 0
                terkategori = False
                namakategori = ''
                kategorifinal = ''
                while it < len(skoring):
                    if it == 0:
                        ranges = range(ra,int(skoring[it])+1)
                    else:
                        ranges = range(ra+1,int(skoring[it])+1)
                    terkategori = int(skor) in ranges
                    namakategori = r_skoring[it]
                    # print(skor, ranges, terkategori, r_skoring[it])
                    ra = int(skoring[it])
                    it += 1
                    if terkategori == True:
                        kategorifinal = namakategori
                        # print('Final Kategori:', kategorifinal)
                sqlkatdelete = "DELETE FROM ts_%s where kdebps='%s' and tahun=%s" % (tabel, kode, tahun)
                engine.execute(sa_text(sqlkatdelete).execution_options(autocommit=True))   
                sqlterkategori = "INSERT INTO ts_%s (kdebps,tahun,skor,kategori) values ('%s',%s,%s,'%s')" % (tabel,kode,tahun,'1',kategorifinal)
                print(sqlterkategori)
                engine.execute(sa_text(sqlterkategori).execution_options(autocommit=True))
            sqlagg = "CREATE OR REPLACE VIEW tv_%s AS SELECT DISTINCT a.kdbbps, a.count, a.kategori, b.tahun FROM (SELECT z.kdbbps, z.count, y.kategori FROM (SELECT t.kdbbps, MAX(t.count) AS count FROM (SELECT DISTINCT kategori,SUBSTRING(kdebps,1,4) AS kdbbps, COUNT(SUBSTRING(kdebps,1,4)) AS count, tahun FROM ts_%s GROUP BY kdbbps, tahun, kategori ORDER BY kdbbps) AS t GROUP BY t.kdbbps ORDER BY t.kdbbps) AS z, (SELECT t.kdbbps, t.kategori, MAX(t.count) AS count FROM (SELECT DISTINCT kategori,SUBSTRING(kdebps,1,4) AS kdbbps, COUNT(SUBSTRING(kdebps,1,4)) AS count, tahun FROM ts_%s GROUP BY kdbbps, tahun, kategori ORDER BY kdbbps) AS t GROUP BY t.kdbbps, t.kategori ORDER BY t.kdbbps)  AS y WHERE z.count = y.count AND y.kdbbps = z.kdbbps ORDER BY z.kdbbps) AS a, ts_%s AS b WHERE a.kdbbps = SUBSTRING(b.kdebps,1,4) ORDER BY a.kdbbps" % (tabel, tabel, tabel, tabel)
            print(sqlagg)
            engine.execute(sqlagg)
            sqlcreategview = "CREATE OR REPLACE VIEW public.tvg_%s AS SELECT a.geom, b.* FROM public.admin_desa a FULL OUTER JOIN public.ts_%s b ON a.kdebps=b.kdebps" % (tabel,tabel)
            engine.execute(sqlcreategview)
            # for item in data:
            #     print(item, data[item][0])
            #     sqlinsert = "INSERT INTO td_%s (indikator, value) values ('%s', '%s')" % (tabel, item, data[item][0])
            #     print(sqlinsert)
            #     engine.execute(sqlinsert)
            # sqlkategori = "INSERT INTO t_kategoribreak (n_variabel, n_break, v_break, v_color) values ('%s',%s,'%s','%s')" % (tabel, 1, jumlah, '#123123')
            # engine.execute(sqlkategori)
            # engine.dispose()        
            # return Response("{OK}", mimetype='application/json') 
            return redirect(url_for('variabel'))

@app.route(config.PREFIX +'/simpanvariabel5', methods=['POST'])
def simpanvariabel5():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        if request.method == 'POST':
            tabel = request.args['tabel']
            tipe = request.args['tipe']
            data = dict(request.form)
            parent = []
            child = []
            subchild = {}
            subsubchild = {}
            jmlchild = 0
            jmlsub = 0
            jmlchildchild = 1
            for item in data:
                jmlsub = jmlsub + 1
                if item.split('_')[0] == 'custkat':
                    tmp = {}
                    tmp['kategori'] = data[item][0]
                    subchild[int(item.split('_')[1])] = tmp
                    jmlchild = jmlchild + 1
            for item1 in data:
                if item1.split('_')[0] == 'subkat':
                    tmp2 = {}
                    tmp3 = {}
                    tmp2['pertanyaan'] = data[item1][0]
                    subchild[int(item1.split('_')[1])][int(item1.split('_')[2])] = tmp2
                    for item2 in data:
                        if item2.split('_')[0] == 'subcuskatqq' and item1.split('_')[1] == item2.split('_')[1]:
                            try:
                                tmp4 = {}
                                tmp4['jawaban'] = data[item2][0]
                                subchild[int(item2.split('_')[1])][int(item2.split('_')[2])][int(item2.split('_')[3])] = tmp4
                            except:
                                pass
                    for item3 in data:
                        if item3.split('_')[0] == 'subcuskatqqs' and item1.split('_')[1] == item3.split('_')[1]:
                            try:
                                tmp5 = {}
                                tmp5['skor'] = data[item3][0]
                                subchild[int(item3.split('_')[1])][int(item3.split('_')[2])][int(item3.split('_')[3])]['skor'] = data[item3][0]
                            except:
                                pass
                while jmlchildchild < jmlsub:
                    # print('Y',item1)
                    jmlchildchild = jmlchildchild+1
            print("SUB2",subchild)
            sqldroptable = "DROP TABLE IF EXISTS public.ti_%s" % (tabel)
            engine.execute(sqldroptable)
            sqlcreatetabel = "CREATE TABLE IF NOT EXISTS public.ti_%s (id BIGSERIAL PRIMARY KEY, kdebps character varying(12), kdepum character varying(12), tahun smallint, skor real)" % (tabel)
            engine.execute(sa_text(sqlcreatetabel).execution_options(autocommit=True))
            # sqlskor = "ALTER TABLE public.ti_%s ADD COLUMN skor real" % (tabel)
            # engine.execute(sqlskor)
            for i in subchild:
                # print('KAT',subchild[i]['kategori'])
                j1 = 0
                while j1 < len(subchild[i])-1:
                    j1 = j1 + 1
                    print('KAT',subchild[i]['kategori'],'PER', subchild[i][j1]['pertanyaan'])
                    sqlkrit = "ALTER TABLE public.ti_%s ADD COLUMN k%s_%s real" % (tabel,i,j1)
                    engine.execute(sa_text(sqlkrit).execution_options(autocommit=True))
                    print(sqlkrit)
                    # print(subchild[i][j1])
                    j2 = 0
                    while j2 < len(subchild[i][j1])-1:
                        j2 = j2 + 1
                        # print(subchild[i][j1])
                        j3 = 0
                        # print('KAT',subchild[i]['kategori'],'PER', subchild[i][j1]['pertanyaan'],'JAW',subchild[i][j1][j2]['jawaban'], 'SKO', subchild[i][j1][j2]['skor'])
                    # print('PER', subchild[i][j1]['pertanyaan'])
                    # print('PER', per)
            sqltruncate = "TRUNCATE TABLE td_%s CASCADE" % (tabel)
            print(sqltruncate)
            engine.execute(sa_text(sqltruncate).execution_options(autocommit=True))
            sqlinsert = "INSERT INTO td_%s (indikator, value) values ('tipe', '%s')" % (tabel, str(tipe))
            print(sqlinsert)
            engine.execute(sqlinsert)
            try:
                sqldel = "DELETE FROM td_%s WHERE indikator='pertanyaan'" % (tabel)
                engine.execute(sa_text(sqldel))
                sql = "INSERT INTO td_%s (indikator, value) values ('pertanyaan','%s')" % (tabel, json.dumps(subchild))
                engine.execute(sa_text(sql))
                print(sql)
            except:
                sql = "INSERT INTO td_%s (indikator, value) values ('pertanyaan','%s')" % (tabel, json.dumps(subchild))
                engine.execute(sa_text(sql))
                print(sql)
            sqlkategori = "INSERT INTO t_kategoribreak (n_variabel, n_break, v_break, v_color) values ('%s',%s,'%s','%s')" % (tabel, 1, 1, '#123123')
            engine.execute(sqlkategori)
            sqltskort = "CREATE TABLE IF NOT EXISTS public.ts_%s (id BIGSERIAL PRIMARY KEY, kdebps character varying(12), kdepum character varying(12), tahun smallint, skor real, kategori text)" % (tabel) 
            engine.execute(sa_text(sqltskort).execution_options(autocommit=True))
            sqlagg = "CREATE OR REPLACE VIEW tv_%s AS SELECT DISTINCT a.kdbbps, a.count, a.kategori, b.tahun FROM (SELECT z.kdbbps, z.count, y.kategori FROM (SELECT t.kdbbps, MAX(t.count) AS count FROM (SELECT DISTINCT kategori,SUBSTRING(kdebps,1,4) AS kdbbps, COUNT(SUBSTRING(kdebps,1,4)) AS count, tahun FROM ts_%s GROUP BY kdbbps, tahun, kategori ORDER BY kdbbps) AS t GROUP BY t.kdbbps ORDER BY t.kdbbps) AS z, (SELECT t.kdbbps, t.kategori, MAX(t.count) AS count FROM (SELECT DISTINCT kategori,SUBSTRING(kdebps,1,4) AS kdbbps, COUNT(SUBSTRING(kdebps,1,4)) AS count, tahun FROM ts_%s GROUP BY kdbbps, tahun, kategori ORDER BY kdbbps) AS t GROUP BY t.kdbbps, t.kategori ORDER BY t.kdbbps)  AS y WHERE z.count = y.count AND y.kdbbps = z.kdbbps ORDER BY z.kdbbps) AS a, ts_%s AS b WHERE a.kdbbps = SUBSTRING(b.kdebps,1,4) ORDER BY a.kdbbps" % (tabel, tabel, tabel, tabel)
            engine.execute(sqlagg)
            sqlcreategview = "CREATE OR REPLACE VIEW public.tvg_%s AS SELECT a.geom, b.* FROM public.admin_desa a FULL OUTER JOIN public.ts_%s b ON a.kdebps=b.kdebps" % (tabel,tabel)
            engine.execute(sqlcreategview)
            # return Response("{OK}", mimetype='application/json') 
            return redirect(url_for('variabel'))

@app.route(config.PREFIX +'/simpanbreak', methods=['POST'])
def simpanbreak():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        if request.method == 'POST':
            tabel = request.args['tabel']
            # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
            indata = dict(request.form)
            print(indata)
            orddata = OrderedDict(sorted(indata.items()))
            isian_break = ''
            isian_warna = ''
            isian_remark = ''
            n_break = 0
            for key, val in orddata.items():
                if key[:5] == 'break':
                    isian_break = isian_break + val[0] + ','
                if key[:4] == 'spec':
                    isian_warna = isian_warna + val[0] + ','
                if key[:3] == 'rem':
                    isian_remark = isian_remark + val[0] + ','
                n_break = n_break + 1
            isian_break = isian_break[:-1]
            isian_warna = isian_warna[:-1]
            isian_remark = isian_remark[:-1]
            sqlbreakclr = "DELETE FROM t_kategoribreak where n_variabel='%s'" % (tabel)
            engine.execute(sa_text(sqlbreakclr).execution_options(autocommit=True))
            sqlbreak = "INSERT INTO t_kategoribreak (n_variabel, n_break, v_break, v_color, r_break) values ('%s',%s,'%s','%s','%s')" % (tabel,n_break/3,isian_break,isian_warna,isian_remark)
            print('SQL: ',sqlbreakclr, sqlbreak)
            engine.execute(sqlbreak)
            # engine.dispose()
            return redirect(url_for('kategori'))
            # return Response("{OK}", mimetype='application/json') 

@app.route(config.PREFIX +'/simpanprofildaerah', methods=['POST'])
def simpanprofildaerah():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        if request.method == 'POST':
            kode = request.args['kode']
            indata = dict(request.form)
            htmldata = indata['htmldata'][0]
            print(indata)
            sqlupdate = "UPDATE t_user SET deskripsi='%s' where username='%s'" % (htmldata.replace("'","''"),kode)
            engine.execute(sa_text(sqlupdate).execution_options(autocommit=True))
            return Response("{OK}", mimetype='application/json') 

@app.route(config.PREFIX +'/profildaerah', methods=['GET'])
def profildaerah():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        kode = request.args['kode']
        deskripsi = t_user.query.with_entities(t_user.username,t_user.desa,t_user.kecamatan,t_user.kabupaten,t_user.provinsi,t_user.deskripsi,t_user.kodepum).filter_by(username=kode).first()
        deskripsi_schema = t_userSchema()
        output = deskripsi_schema.dump(deskripsi)
        jsonoutput = json.dumps(output)
        return Response(jsonoutput, mimetype='application/json')

# @app.route(config.PREFIX +'/profiladmin',methods=['GET'])
# def profiladmin():
#     kode = request.args['kode']

@app.route(config.PREFIX +'/simpanindikator', methods=['POST'])
def simpanindikator():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        if request.method == 'POST':
            tahun = request.args['tahun']
            tabel = request.args['tabel']
            kode = request.args['kode']
            model = 1
            try:
                model = request.args['model']
                indata = dict(request.form)
                print("MODEL",model,indata)
            except:
                pass
            # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
            indata = dict(request.form)
            orddata = OrderedDict(sorted(indata.items()))
            kolom = 'kdebps,tahun,'
            isian = "'%s',%s," % (kode,tahun)
            flag = ''
            skor = 0
            # print(indata)
            if model == '1':
                for key, val in orddata.items():
                    if val[0] == '0':
                        flag = 'False'
                        skor = skor + 0
                    if val[0] == '1':
                        flag = 'True'
                        skor = skor + 1
                    kolom = kolom + key + ','
                    isian = isian + flag + ',' 
            if model == '2':
                minimumskorarr = []
                for key, val in orddata.items():
                    flag = val[0]
                    kolom = kolom + key + ','
                    isian = isian + flag + ','
                    minimumskorarr.append(int(flag))       
                skor = min(minimumskorarr)          
            kolom = kolom[:-1]
            isian = isian[:-1]
            print('ITEM', tabel, kolom, isian, orddata)
            try:
                sqldelete = "DELETE FROM ti_%s where kdebps='%s' and tahun=%s" % (tabel, kode, tahun)
                engine.execute(sa_text(sqldelete).execution_options(autocommit=True))
                if model == '2':
                    htmldata = indata
                    print(htmldata)
                if model == '4':
                    # isian = isian + "'" + base64.b64encode(str.encode(indata['htmldata'][0])).decode() + "'"
                    htmldata = indata['htmldata'][0]
                    print(isian)
                    isian = isian + ",'" + htmldata.replace("'","") + "'"
                    print(htmldata)
                    sqlinsert = "INSERT INTO ti_%s (%s,htmldata,skor) values (%s,%s)" % (tabel, kolom, isian, skor)
                    print("X",sqlinsert)
                    engine.execute(sa_text(sqlinsert))
                if model =='5':
                    print(orddata)
                    kolom = 'kdebps,tahun,'
                    isian = "'%s',%s," % (kode,tahun)
                    for i in orddata:
                        skor = skor + int(orddata[i][0])
                        kolom = kolom + i + ','
                        isian = isian + orddata[i][0] + ','
                    kolom = kolom[:-1]
                    isian = isian[:-1]
                    print('ITEM',kolom, isian, skor)
                    sqlinsert = "INSERT INTO ti_%s (%s,skor) values (%s,%s)" % (tabel, kolom, isian, skor)
                    engine.execute(sa_text(sqlinsert))
                    print("X",sqlinsert)
                if model == '1' or model == '3':
                    sqlinsert = "INSERT INTO ti_%s (%s,skor) values (%s,%s)" % (tabel, kolom, isian, skor)
                    engine.execute(sa_text(sqlinsert))
                    print("Y",sqlinsert)
            except:     
                if model == '2':
                    htmldata = indata
                    print(htmldata)
                if model == '4':
                    htmldata = indata['htmldata'][0]
                    isian = isian + "'" + htmldata.replace("'","") + "'"
                sqlinsert = "INSERT INTO ti_%s (%s,skor) values (%s,%s)" % (tabel, kolom, isian, skor)
                print("SQL",sqlinsert)
                engine.execute(sa_text(sqlinsert))
            # print('SQL',sqlinsert)
            sqltskort = "CREATE TABLE IF NOT EXISTS public.ts_%s (id BIGSERIAL PRIMARY KEY, kdebps character varying(12), kdepum character varying(12), tahun smallint, skor real, kategori text)" % (tabel) 
            engine.execute(sa_text(sqltskort).execution_options(autocommit=True))
            kategoribreak = t_kategoribreak.query.filter_by(n_variabel=tabel).first()
            kategoribreak_schema = t_kategoribreakSchema()
            output = kategoribreak_schema.dump(kategoribreak).data
            print(output)
            skoring = output['v_break'].split(',')
            r_skoring = output['r_break'].split(',')
            ranges = range(0)
            it = 0
            ra = 0
            terkategori = False
            namakategori = ''
            kategorifinal = ''
            while it < len(skoring):
                if it == 0:
                    ranges = range(ra,int(skoring[it])+1)
                else:
                    ranges = range(ra+1,int(skoring[it])+1)
                terkategori = int(skor) in ranges
                namakategori = r_skoring[it]
                print(skor, ranges, terkategori, r_skoring[it])
                ra = int(skoring[it])
                it += 1
                if terkategori == True:
                    kategorifinal = namakategori
                    print('Final Kategori:', kategorifinal)
            sqlkatdelete = "DELETE FROM ts_%s where kdebps='%s' and tahun=%s" % (tabel, kode, tahun)
            engine.execute(sa_text(sqlkatdelete).execution_options(autocommit=True))   
            if model == '4':
                sqlterkategori = "INSERT INTO ts_%s (kdebps,tahun,skor,kategori) values ('%s',%s,%s,'%s')" % (tabel,kode,tahun,skor,kategorifinal)
            else:             
                sqlterkategori = "INSERT INTO ts_%s (kdebps,tahun,skor,kategori) values ('%s',%s,%s,'%s')" % (tabel,kode,tahun,'1',kategorifinal)
            print(sqlterkategori)
            engine.execute(sqlterkategori)
            # engine.dispose()
            return Response("{OK}", mimetype='application/json') 
            return redirect(url_for('isivariabel'))
            # return Response(sqlinsert, mimetype='application/json') 

@app.route(config.PREFIX +'/htmlprofile', methods=['GET'])
def htmlprofile():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        tabel = request.args['tabel']
        tahun = request.args['tahun']
        kode = request.args['kode']
        # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
        sqlgethtml = "SELECT * from ti_%s WHERE tahun=%s AND kdebps='%s'" % (tabel,tahun,kode)
        print(sqlgethtml)
        result = engine.execute(sa_text(sqlgethtml))
        resultkey = engine.execute(sa_text(sqlgethtml)).keys()
        row = result.fetchone()
        print(row)
        output = {}
        try:
            for key, val in zip(resultkey,row):
                output[key] = val
        except:
            pass
        try:
            jsonoutput = json.dumps(output['htmldata'])
        except:
            jsonoutput = json.dumps('')
        # engine.dispose()
        return Response(jsonoutput, mimetype='application/json')

@app.route(config.PREFIX +'/grafmodel', methods=['GET'])
def grafmodel():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        tabel = request.args['tabel']
        tahun = request.args['tahun']
        kabkode = request.args['kode']
        # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
        sqlgetgraph = "SELECT kategori, COUNT(kategori) as jumlah from ts_%s WHERE tahun=%s AND kdebps LIKE '%s%%' GROUP BY kategori ORDER BY kategori" % (tabel,tahun,kabkode)
        print(sqlgetgraph)
        result = engine.execute(sa_text(sqlgetgraph))
        output = []
        inner = {}
        for item in result:
            inner = {}
            inner['kategori'] = item[0]
            inner['jumlah'] = item[1]
            print(inner)
            output.append(inner)
        # engine.dispose()
        jsonoutput = json.dumps(output)
        return Response(jsonoutput, mimetype='application/json')

@app.route(config.PREFIX +'/grafmodel1c3', methods=['GET'])
def grafmodel1c3():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        tabel = request.args['tabel']
        tahun = request.args['tahun']
        kabkode = request.args['kode']
        # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
        sqlgetgraph = "SELECT kategori, COUNT(kategori) as jumlah from ts_%s WHERE tahun=%s AND kdebps LIKE '%s%%' GROUP BY kategori ORDER BY kategori" % (tabel,tahun,kabkode)
        print(sqlgetgraph)
        result = engine.execute(sa_text(sqlgetgraph))
        output = []
        for item in result:
            inner = [item[0],item[1]]
            output.append(inner)
        # engine.dispose()
        jsonoutput = json.dumps(output)
        return Response(jsonoutput, mimetype='application/json')

@app.route(config.PREFIX +'/grafmodel1c3agg', methods=['GET'])
def grafmodel1c3agg():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        tabel = request.args['tabel']
        tahun = request.args['tahun']
        # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
        sqlgetgraph = "SELECT kategori, COUNT(skor) as jumlah from ts_%s WHERE tahun=%s GROUP BY kategori ORDER BY kategori" % (tabel,tahun)
        # sqlgetgraph = "SELECT kategori, SUM(count) as jumlah from tv_%s WHERE tahun=%s GROUP BY kategori ORDER BY kategori" % (tabel,tahun)
        print(sqlgetgraph)
        result = engine.execute(sa_text(sqlgetgraph))
        output = []
        for item in result:
            jml = str(item[1])
            inner = [item[0],jml]
            output.append(inner)
        # engine.dispose()
        jsonoutput = json.dumps(output)
        return Response(jsonoutput, mimetype='application/json')

@app.route(config.PREFIX +'/mapcolor', methods=['GET'])
def mapcolor():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        tabel = request.args['tabel']
        tahun = request.args['tahun']
        kategoribreak = t_kategoribreak.query.filter_by(n_variabel=tabel).first()
        kategoribreak_schema = t_kategoribreakSchema()
        kategoribreakdata = kategoribreak_schema.dump(kategoribreak).data
        # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)    
        sqlgetcolor = "SELECT kdebps, kategori as jumlah from ts_%s WHERE tahun=%s GROUP BY kdebps, kategori ORDER BY kdebps" % (tabel,tahun)
        print(sqlgetcolor)
        result = engine.execute(sa_text(sqlgetcolor))
        output = {}
        try:
            colors = kategoribreakdata['v_color'].split(',')
            breaks = kategoribreakdata['r_break'].split(',')
            print(colors)
            for item in result:
                inner = {}
                for i,j in zip(breaks, colors):
                    if i == item[1]:
                        output[item[0]] = j
                # output.append(inner) 
            print(kategoribreakdata)
            # engine.dispose()
        except:
            pass
        jsonoutput = json.dumps(output)
        return Response(jsonoutput, mimetype='application/json')

@app.route(config.PREFIX +'/mapcoloragg', methods=['GET'])
def mapcoloragg():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        tabel = request.args['tabel']
        tahun = request.args['tahun']
        kategoribreak = t_kategoribreak.query.filter_by(n_variabel=tabel).first()
        kategoribreak_schema = t_kategoribreakSchema()
        kategoribreakdata = kategoribreak_schema.dump(kategoribreak).data
        # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)    
        sqlgetcolor = "SELECT kdbbps, kategori as jumlah from tv_%s WHERE tahun=%s GROUP BY kdbbps, kategori ORDER BY kdbbps" % (tabel,tahun)
        print(sqlgetcolor)
        result = engine.execute(sa_text(sqlgetcolor))
        output = {}
        colors = kategoribreakdata['v_color'].split(',')
        breaks = kategoribreakdata['r_break'].split(',')
        print(colors)
        for item in result:
            inner = {}
            for i,j in zip(breaks, colors):
                if i == item[1]:
                    output[item[0]] = j
            # output.append(inner) 
        print(kategoribreakdata)
        # engine.dispose()
        jsonoutput = json.dumps(output)
        return Response(jsonoutput, mimetype='application/json')

@app.route(config.PREFIX +'/kategori')
def kategori():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        print(landing_judul.config_desc)
        variabel = t_variabel.query.order_by(t_variabel.id.asc()).all()
        kategoribreak = t_kategoribreak.query.order_by(t_kategoribreak.id.asc()).all()
        return render_template('dashboard/kategori.htm', landing_judul=landing_judul.config_desc,landing_logo=landing_logo.config_desc, variabel=variabel, kategori=kategori,title=landing_judul.config_desc)

@app.route(config.PREFIX +'/kategoridesc')
def kategoridesc():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        variabel = request.args.get('variabel')
        kategoridesc = t_kategoribreak.query.filter_by(n_variabel=variabel).first()
        kategoridesc_schema = t_kategoribreakSchema()
        output = kategoridesc_schema.dump(kategoridesc)
        jsonoutput = json.dumps(output)
        return Response(jsonoutput, mimetype='application/json')

@app.route(config.PREFIX +'/simpantemplateisian', methods=['POST'])
def simpantemplateisian():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        if request.method == 'POST':
            tabel = request.args['tabel']
            # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
            result = engine.execute('select * from td_' + tabel + ' order by id')
            output = {}
            for item in result:
                print(item.indikator, item.value)
                output[item.indikator] = item.value
            wb = Workbook()
            print(output)
            fl = tabel + '_' + randomstr_generator() + '.xlsx'
            dest_filename = config.APP_ROOT + 'app/static/data/ex_templates/' + fl
            ws1 = wb.active
            ws1.title = 'Isian Data'
            ws2 = wb.create_sheet(title="Keterangan - Pejelasan")
            wb.save(filename = dest_filename)
            # engine.dispose()
            return Response("{filename:%s}" % (fl), mimetype='application/json') 

@app.route(config.PREFIX +'/getdesainkab', methods=['GET'])
def getdesainkab():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        kode = request.args['kode']
        massout = []
        # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
        result = engine.execute("select * from t_user where left(username,4) like '" + kode + "'")
        for item in result:
            output = {}
            output['kode'] = item.username
            output['desa'] = item.desa
            massout.append(output)
        jsonoutput = json.dumps(massout)
        print(output)
        # engine.dispose()
        return Response(jsonoutput, mimetype='application/json')

@app.route(config.PREFIX +'/getjawaban', methods=['GET'])
def getjawaban():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        kode = request.args['kode']
        tabel = request.args['tabel']
        tahun = request.args['tahun']
        # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
        tipepertanyaansql = "SELECT indikator, value FROM td_%s WHERE indikator = 'tipe'" % (tabel)
        pertanyaansql = "SELECT indikator, value FROM td_%s WHERE indikator != 'tipe'" % (tabel)
        jawabansql = "SELECT * FROM ti_%s WHERE kdebps='%s' AND tahun=%s" % (tabel,kode,tahun)
        restipepertanyaansql = engine.execute(tipepertanyaansql)
        respertanyaansql = engine.execute(pertanyaansql)
        resjawabansql = engine.execute(jawabansql)
        resjawabansqlkey = engine.execute(jawabansql).keys()
        for item in restipepertanyaansql:
            # print("TIPE", item)
            pass
        for item in respertanyaansql:
            # print("PERTANYAAN", item)
            pass
        output = {}
        for item in resjawabansql:
            inner = {}
            for i, o in zip(item, resjawabansqlkey):
                print("JAWABAN", o, i)
                if(o == "id" and o == "kdebps" and o == "kdepum" and o == "tahun" and o == "skor"):
                    pass
                else:
                    output[o] = i     
        jsonoutput = json.dumps(output)
        # engine.dispose()
        return Response(jsonoutput, mimetype='application/json')

@app.route(config.PREFIX +'/profileagg', methods=['GET'])
def profileagg():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        # kode = request.args['kode']
        tabel = request.args['tabel']
        tahun = request.args['tahun']
        # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
        sqltd = "SELECT indikator, value FROM td_%s WHERE indikator='tipe'" % (tabel)
        sql = "SELECT * FROM tv_%s WHERE tahun=%s" % (tabel, tahun)
        result = engine.execute(sql)
        resultkey = engine.execute(sql).keys()
        restd = engine.execute(sqltd)
        output = []
        rowtd = restd.fetchone()
        for item in result:
            inner = {}
            for key, val in zip(resultkey,item):
                inner[key] = val
                inner[rowtd.indikator] = rowtd.value
            output.append(inner)
        jsonoutput = json.dumps(output)
        # engine.dispose()
        return Response(jsonoutput, mimetype='application/json')

@app.route(config.PREFIX +'/penduduk', methods=['GET'])
def penduduk():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        kode = request.args['kode']
        if kode == 'undefined':
            sql = "SELECT sum(b.jumlah_kk) as jumlah_kk, sum(b.jml_laki) as jml_laki, sum(b.jml_perempuan) as jml_perempuan, sum(b.kepadatan1) as kepadatan1,sum(b.u0_4) as u0_4, sum(b.u5_9) as u5_9, sum(b.u10_14) as u10_14, sum(b.u15_19) as u15_19, sum(b.u20_24) as u20_24, sum(b.u25_29) as u25_29, sum(b.u30_34) as u30_34, sum(b.u35_39) as u35_39, sum(b.u40_44) as u40_44,sum(b.u45_49) as u45_49, sum(b.u50_54) as u50_54, sum(b.u55_59) as u55_59, sum(b.u60_64) as u60_64, sum(b.u65_69) as u65_69, sum(b.u70_74) as u70_74, sum(b.u75) as u75 FROM tb_penduduk_all b WHERE substring(b.kdebps,1,2) != 0::text"
            result = engine.execute(sql)
            resultkey = engine.execute(sql).keys()
            output = []
            for item in result.fetchall():
                inner = {}
                for key, val in zip(resultkey,item):
                    inner[key] = str(val)
                output.append(inner)
        elif len(kode) == 4:
            sql = "SELECT substring(a.kdebps,1,4) as kdbbps,sum(b.jumlah_kk) as jumlah_kk, sum(b.jml_laki) as jml_laki, sum(b.jml_perempuan) as jml_perempuan, sum(b.kepadatan1) as kepadatan1,sum(b.u0_4) as u0_4, sum(b.u5_9) as u5_9, sum(b.u10_14) as u10_14, sum(b.u15_19) as u15_19, sum(b.u20_24) as u20_24, sum(b.u25_29) as u25_29, sum(b.u30_34) as u30_34, sum(b.u35_39) as u35_39, sum(b.u40_44) as u40_44,sum(b.u45_49) as u45_49, sum(b.u50_54) as u50_54, sum(b.u55_59) as u55_59, sum(b.u60_64) as u60_64, sum(b.u65_69) as u65_69, sum(b.u70_74) as u70_74, sum(b.u75) as u75 FROM tb_penduduk_all a, tb_penduduk_all b WHERE substring(a.kdebps,1,4) = substring(b.kdebps,1,4) AND substring(b.kdebps,1,4) != 0::text and substring(b.kdebps,1,4)='%s' GROUP BY substring(a.kdebps,1,4)" % (kode)
            result = engine.execute(sql)
            resultkey = engine.execute(sql).keys()
            output = []
            for item in result.fetchall():
                inner = {}
                for key, val in zip(resultkey,item):
                    inner[key] = str(val)
                output.append(inner)
        elif len(kode) == 6:
            sql = "SELECT substring(a.kdebps,1,4) as kdbbps,sum(b.jumlah_kk) as jumlah_kk, sum(b.jml_laki) as jml_laki, sum(b.jml_perempuan) as jml_perempuan, sum(b.kepadatan1) as kepadatan1,sum(b.u0_4) as u0_4, sum(b.u5_9) as u5_9, sum(b.u10_14) as u10_14, sum(b.u15_19) as u15_19, sum(b.u20_24) as u20_24, sum(b.u25_29) as u25_29, sum(b.u30_34) as u30_34, sum(b.u35_39) as u35_39, sum(b.u40_44) as u40_44,sum(b.u45_49) as u45_49, sum(b.u50_54) as u50_54, sum(b.u55_59) as u55_59, sum(b.u60_64) as u60_64, sum(b.u65_69) as u65_69, sum(b.u70_74) as u70_74, sum(b.u75) as u75 FROM tb_penduduk_all a, tb_penduduk_all b WHERE substring(a.kdebps,1,4) = substring(b.kdebps,1,4) AND substring(b.kdebps,1,4) != 0::text and substring(b.kdebps,1,4)='%s' GROUP BY substring(a.kdebps,1,4)" % (kode)
            result = engine.execute(sql)
            resultkey = engine.execute(sql).keys()
            output = []
            for item in result.fetchall():
                inner = {}
                for key, val in zip(resultkey,item):
                    inner[key] = str(val)
                output.append(inner)
        else:
            # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
            sql = "SELECT * FROM tb_penduduk_all WHERE kdebps='%s'" % (kode)
            result = engine.execute(sql)
            resultkey = engine.execute(sql).keys()
            output = {}
            for item in result:
                for key, val in zip(resultkey,item):
                    output[key] = str(val)
        print(output)
        jsonoutput = json.dumps(output)
        # engine.dispose()
        return Response(jsonoutput, mimetype='application/json')

@app.route(config.PREFIX +'/pendidikan', methods=['GET'])
def pendidikan():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        kode = request.args['kode']
        # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
        sql = "SELECT * FROM tb_pendidikan WHERE kdebps='%s'" % (kode)
        result = engine.execute(sql)
        resultkey = engine.execute(sql).keys()
        output = {}
        for item in result:
            for key, val in zip(resultkey,item):
                output[key] = val
        jsonoutput = json.dumps(output)
        # engine.dispose()
        return Response(jsonoutput, mimetype='application/json')

@app.route(config.PREFIX +'/radar', methods=['GET'])
def radar():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        tabel = request.args['tabel']
        kode = request.args['kode']
        tahun = request.args['tahun']    
        sql1 = "SELECT * FROM ti_%s WHERE kdebps='%s' AND tahun=%s" % (tabel,kode, tahun)
        sql2 = "SELECT value FROM td_%s WHERE indikator='pertanyaan'" % (tabel)
        res1 = engine.execute(sql1)
        key1 = engine.execute(sql1).keys()
        res2 = engine.execute(sql2)
        output = {}
        kkeys = {}
        kate = []
        katkat = []
        nilai = []
        for item3 in res2:
            kat = json.loads(item3[0])
            # print(kat)
            for item4 in kat:
                # print('I4', kat[item4])
                kate.append('k' + str(item4))
                katkat.append(kat[item4]['kategori'])
        # print(kate)
        for item1 in res1:
            for key, val in zip(key1,item1):
                if key[:1] == 'k' and key != 'kdebps' and key !='kdepum':
                    kkeys[key] = val
        for item2 in kate:
            tempnilai = 0
            for item5 in kkeys:
                if item5[:2] == item2:
                    tempnilai = tempnilai + kkeys[item5]
            nilai.append(tempnilai)
            # print(item5)
            # print(item2)
        # print(kate,nilai)
        output['keterangan'] = katkat
        output['katnum'] = kate
        output['nilai'] = nilai
        jsonoutput = json.dumps(output)
        # engine.dispose()
        return Response(jsonoutput, mimetype='application/json')

@app.route(config.PREFIX +'/pesisir', methods=['GET'])
def pesisir():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        kode = request.args['kode']
        tahun = request.args['tahun']
        # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
        sql = "SELECT * FROM ti_t_pesisir WHERE kdebps='%s' AND tahun=%s" % (kode, tahun)
        result = engine.execute(sql)
        resultkey = engine.execute(sql).keys()
        output = []
        for item in result:
            for key, val in zip(resultkey,item):
                if key.split('_')[0] == 'asp':
                    # output[key] = val
                    print(key,val)
                    output.append(val)
        jsonoutput = json.dumps(output)
        # engine.dispose()
        return Response(jsonoutput, mimetype='application/json')

@app.route(config.PREFIX +'/prodes', methods=['GET'])
def prodes():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        # tabel = request.args['tabel']
        # tahun = request.args['tahun']
        deskode = request.args['kode']
        # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
        sqlgetdemografi = "SELECT idbpsdesa,desa,kecamatan,kabupaten,provinsi from admin_desa WHERE idbpsdesa='%s'" % (deskode)
        print(sqlgetdemografi)
        result = engine.execute(sqlgetdemografi)
        resultkey = engine.execute(sqlgetdemografi).keys()
        output = []
        inner = {}
        for item in result:
            inner = {}
            for key, val in zip(resultkey,item):
                inner[key] = val
                # print(key,val)
            output.append(inner)
        jsonoutput = json.dumps(output)
        # engine.dispose()
        return Response(jsonoutput, mimetype='application/json')

@app.route(config.PREFIX +'/prokab', methods=['GET'])
def prokab():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        kabkode = request.args['kode']
        sqlgetdemografi = "SELECT distinct kdbbps,kabupaten,provinsi from admin_desa WHERE kdbbps='%s' group by kdbbps,kabupaten,provinsi" % (kabkode)
        print(sqlgetdemografi)
        result = engine.execute(sqlgetdemografi)
        resultkey = engine.execute(sqlgetdemografi).keys()
        output = []
        inner = {}
        for item in result:
            inner = {}
            for key, val in zip(resultkey,item):
                inner[key] = val
            output.append(inner)
        jsonoutput = json.dumps(output)
        return Response(jsonoutput, mimetype='application/json')

@app.route(config.PREFIX +'/fdesa', methods=['GET'])
def fdesa():
    rows = t_user.query.order_by(t_user.username).all()
    output = []
    inner = {}
    for item in rows:
        inner = {}
        inner['category'] = item.kabupaten
        inner['label'] = item.desa
        inner['kodebps'] = item.username
        output.append(inner)
    jsonoutput = json.dumps(output)
    return Response(jsonoutput, mimetype='application/json')

@app.route(config.PREFIX +'/eksportmpl', methods=['GET'])
def eksportmpl():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        if request.method == 'GET':
            tabel = request.args['tabel']
            tipe = request.args['tipe']
            sql1 = "SELECT indikator, value FROM td_%s" % (tabel)
            result1 = engine.execute(sql1)
            isian = []
            wb = Workbook()
            ws = wb.active
            ws.title = "Keterangan"
            ws1 = wb.create_sheet("Isian",1)
            ws1.title = "Isian"
            if tipe == '1':
                ws['A1'] = 'KETERANGAN ISIAN'
                ws['A2'] = '----------------'
                ws['A3'] = 'Berikut adalah keterangan kolom isian.'
                rws = 5
                ws1['A1'] = 'kdebps'
                ws1['B1'] = 'tahun'
                irow = 1
                icol = 3
                for item in result1:
                    if item[0] != 'tipe':
                        ws['A' + str(rws)] = item[0]
                        ws['B' + str(rws)] = item[1]
                        rws = rws + 1
                        ws1.cell(row=irow, column=icol, value=item[0])
                        icol = icol+1
            if tipe == '2':
                ws['A1'] = 'KETERANGAN ISIAN'
                ws['A2'] = '----------------'
                ws['A3'] = 'Berikut adalah keterangan kolom isian.'
                rws = 5
                rwp = 6
                iket = 1
                irow = 1
                icol = 3
                ws1['A1'] = 'kdebps'
                ws1['B1'] = 'tahun'
                for item in result1:
                    if item[0] == 'pertanyaan':
                        qper = ast.literal_eval(item[1])
                        for qitem in qper:
                            ws['A' + str(rws)] = 'pilgan_' + str(qitem)
                            ws['A' + str(rwp)] = 'Jawaban:'
                            ws['B' + str(rws)] = qper[(qitem)]['1']
                            qlen = len(qper[qitem])
                            jawaban = ','
                            for qq in qper[(qitem)]:
                                print(qq)
                                if qq != '1':
                                    jawaban = jawaban + str(int(qq)-1) + '. ' + qper[(qitem)][qq] + ', '
                            print(jawaban)
                            jawaban = jawaban[1:-2]
                            ws['B' + str(rwp)] = jawaban
                            ws1.cell(row=irow, column=icol, value='pilgan_' + str(qitem))
                            rws = rws + 2
                            rwp = rwp + 2 
                            iket = iket + 1 
                            icol = icol+1
            if tipe == '5':
                ws['A1'] = 'KETERANGAN ISIAN'
                ws['A2'] = '----------------'
                ws['A3'] = 'Berikut adalah keterangan kolom isian.'
                rws = 5
                rwp = 6
                iket = 1
                irow = 1
                icol = 3
                ws1['A1'] = 'kdebps'
                ws1['B1'] = 'tahun'
                for item in result1:
                    if item[0] == 'pertanyaan':
                        uqper = ast.literal_eval(item[1])
                        qper = sorted(uqper.items(), key=operator.itemgetter(0))
                        # print('ITEM',qper)
                        iter1 = 0
                        peri = 0
                        while iter1 < len(qper):
                            ws['A' + str(rws+peri)] = qper[iter1][1]['kategori']
                            iter2 = 1
                            # while iter2 < len(qper[iter1][1]['1']):
                            while iter2 < len(qper[iter1][1]):
                                ws['B' + str(rws+peri)] = 'k' + str(qper[iter1][0]) + '_' + str(iter2)
                                ws1.cell(row=irow, column=icol, value='k' + str(qper[iter1][0]) + '_' + str(iter2))
                                icol = icol+1
                                # print(qper[iter1][1])
                                iter3 = 1
                                while iter3 < len(qper[iter1][1]['1']):
                                    # print(qper[iter1][1][str(iter3)])
                                    ws['C' + str(rws+peri)] = str(qper[iter1][1][str(iter2)][str(iter3)]['skor'])
                                    ws['D' + str(rws+peri)] = str(qper[iter1][1][str(iter2)][str(iter3)]['jawaban'])
                                    iter3 = iter3 + 1
                                    peri = peri + 1
                                iter2 = iter2 + 1
                            iter1 = iter1 + 1
            vwb=save_virtual_workbook(wb)
            output = make_response(vwb)
            rndfile = randomstr_generator()
            wb.save(APP_ROOT + 'app/static/data/tmp/'+ rndfile +'.xlsx')
            return send_from_directory(directory=APP_ROOT + 'app/static/data/tmp/', filename=rndfile +'.xlsx')

@app.route(config.PREFIX +'/impordata', methods=['POST'])
def impordata():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        if request.method == 'POST':
            tipe = request.args['tipe']
            tabel = request.args['tabel']
            if 'qqfile' not in request.files:
                return Response('{"success": false}', mimetype='application/json') 
            file = request.files['qqfile']
            if file.filename == '':
                return Response('{"success": false}', mimetype='application/json') 
            if file:
                filename = secure_filename(file.filename)
                file.save(os.path.join(APP_ROOT + 'app/static/data/tmp/', filename))
                wb = load_workbook(APP_ROOT + 'app/static/data/tmp/' + filename)
                ws = wb['Isian']
                fs = list(ws.rows)[0]
                fsrows = ws.max_row 
                sdrow = 1
                kolom = ''
                isian = ''
                kode = ''
                tahun = ''
                skor = 0
                for kol in fs:
                    kolom = kolom + kol.value + ', '
                kolom = kolom[:-2]
                # print(kolom)
                if tipe == '1':
                    while sdrow < fsrows:
                        # print(list(ws.rows)[sdrow])
                        kolval = ''
                        skor = 0
                        koln = 1
                        for kol in list(ws.rows)[sdrow]:
                            # print(kol)
                            try:
                                if int(kol.value) == 1:
                                    kolval = kolval+'True'+ ', '
                                    skor = skor + 1
                                else:
                                    kolval = kolval+kol.value+ ', '
                            except:
                                if kol.value is None or int(kol.value) == 0:
                                    kolval = kolval+'False'+ ', '
                                else:
                                    kolval = kolval+str(kol.value)+ ', '
                                    if koln == 1:
                                        kode = str(kol.value)
                                    if koln == 2:
                                        tahun = str(kol.value)
                            koln = koln + 1
                        sdrow = sdrow + 1
                        kolval = kolval[:-2]
                        impordatacalc(tipe,tabel,kolom,kolval,skor,kode,tahun)
                    # print(kolval)
                if tipe == '2':
                    while sdrow < fsrows:
                        # print(list(ws.rows)[sdrow])
                        kolval = ''
                        skor = 0
                        koln = 1                
                        minimumskorarr = []   
                        for kol in list(ws.rows)[sdrow]:
                            print(kol, kol.value)
                            kolval = kolval+str(kol.value)+ ', '
                            if koln == 1:
                                kode = str(kol.value)
                            if koln == 2:
                                tahun = str(kol.value)     
                            if koln > 2:         
                                print('2',kol, kol.value)  
                                minimumskorarr.append(int(kol.value))                  
                            # skor = skor + int(kol.value) 
                            koln = koln + 1
                        skor = min(minimumskorarr)  
                        sdrow = sdrow + 1
                        kolval = kolval[:-2]
                        impordatacalc(tipe,tabel,kolom,kolval,skor,kode,tahun)
                    print(kolval)
                if tipe == '5':
                    while sdrow < fsrows:
                        # print(list(ws.rows)[sdrow])
                        kolval = ''
                        skor = 0
                        koln = 1
                        for kol in list(ws.rows)[sdrow]:
                            print(kol, kol.value)
                            kolval = kolval+str(kol.value)+ ', '
                            if koln == 1:
                                kode = str(kol.value)
                            if koln == 2:
                                tahun = str(kol.value)     
                            if koln > 2:                         
                                skor = skor + int(kol.value) 
                            koln = koln + 1
                        sdrow = sdrow + 1
                        kolval = kolval[:-2]
                        impordatacalc(tipe,tabel,kolom,kolval,skor,kode,tahun)
                    print(kolval)
            return Response('{"success": true}', mimetype='application/json') 

@app.route(config.PREFIX +'/eksportmpldasar', methods=['GET'])
def eksportmpldasar():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        if request.method == 'GET':
            pass

def impordatacalc(tipe,tabel,kolom,isian,skor,kode,tahun):
    try:
        sqldelete = "DELETE FROM ti_%s where kdebps='%s' and tahun=%s" % (tabel, kode, tahun)
        print(sqldelete)
        engine.execute(sa_text(sqldelete).execution_options(autocommit=True))
        if tipe == '2':
            sqlinsert = "INSERT INTO ti_%s (%s,skor) values (%s,%s)" % (tabel, kolom, isian, skor)
            engine.execute(sa_text(sqlinsert))
            print("X",sqlinsert)
        if tipe == '5':
            sqlinsert = "INSERT INTO ti_%s (%s,skor) values (%s,%s)" % (tabel, kolom, isian, skor)
            engine.execute(sa_text(sqlinsert))
            print("X",sqlinsert)
        else:
            sqlinsert = "INSERT INTO ti_%s (%s,skor) values (%s,%s)" % (tabel, kolom, isian, skor)
            engine.execute(sa_text(sqlinsert))
            print("Y",sqlinsert)
    except:     
        if tipe == '2':
            sqlinsert = "INSERT INTO ti_%s (%s,skor) values (%s,%s)" % (tabel, kolom, isian, skor)
            engine.execute(sa_text(sqlinsert))
            print("X",sqlinsert)
        sqlinsert = "INSERT INTO ti_%s (%s,skor) values (%s,%s)" % (tabel, kolom, isian, skor)
        print("SQL",sqlinsert)
        engine.execute(sa_text(sqlinsert))
    print('SQL',sqlinsert)
    sqltskort = "CREATE TABLE IF NOT EXISTS public.ts_%s (id BIGSERIAL PRIMARY KEY, kdebps character varying(12), kdepum character varying(12), tahun smallint, skor real, kategori text)" % (tabel) 
    engine.execute(sa_text(sqltskort).execution_options(autocommit=True))
    kategoribreak = t_kategoribreak.query.filter_by(n_variabel=tabel).first()
    kategoribreak_schema = t_kategoribreakSchema()
    output = kategoribreak_schema.dump(kategoribreak).data
    print(output)
    skoring = output['v_break'].split(',')
    r_skoring = output['r_break'].split(',')
    ranges = range(0)
    it = 0
    ra = 0
    terkategori = False
    namakategori = ''
    kategorifinal = ''
    while it < len(skoring):
        if it == 0:
            ranges = range(ra,int(skoring[it])+1)
        else:
            ranges = range(ra+1,int(skoring[it])+1)
        terkategori = int(skor) in ranges
        namakategori = r_skoring[it]
        print(skor, ranges, terkategori, r_skoring[it])
        ra = int(skoring[it])
        it += 1
        if terkategori == True:
            kategorifinal = namakategori
            print('Final Kategori:', kategorifinal)
    sqlkatdelete = "DELETE FROM ts_%s where kdebps='%s' and tahun=%s" % (tabel, kode, tahun)
    engine.execute(sa_text(sqlkatdelete).execution_options(autocommit=True))   
    if tipe == '4':
        sqlterkategori = "INSERT INTO ts_%s (kdebps,tahun,skor,kategori) values ('%s',%s,%s,'%s')" % (tabel,kode,tahun,skor,kategorifinal)
    else:             
        sqlterkategori = "INSERT INTO ts_%s (kdebps,tahun,skor,kategori) values ('%s',%s,%s,'%s')" % (tabel,kode,tahun,skor,kategorifinal)
    print(sqlterkategori)
    engine.execute(sqlterkategori)
    # engine.dispose()

@app.route(config.PREFIX +'/isiintervensi')
def isiintervensi():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        variabel = t_variabel.query.order_by(t_variabel.id.asc()).all()
        # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
        provinsi = "select distinct provinsi, left(username, 2) as username from t_user where provinsi != 'ADMIN' and left(username, 2) !='' and username !='0' order by left(username, 2)" 
        kabupaten = "select distinct kabupaten, left(username, 4) as username from t_user where provinsi != 'ADMIN' and left(username, 4) !='' and kabupaten !='-' and username !='0' order by kabupaten"
        desa = "select distinct desa, username from t_user where provinsi != 'ADMIN' order by desa"
        l_provinsi = engine.execute(provinsi)
        l_kabupaten = engine.execute(kabupaten)
        l_desa = engine.execute(desa)
        # engine.dispose()
        return render_template('dashboard/isiintervensi.htm', landing_logo=landing_logo.config_desc, landing_judul=landing_judul.config_desc,variabel=variabel, provinsi=l_provinsi, kabupaten=l_kabupaten, desa=l_desa,title=landing_judul.config_desc)

@app.route(config.PREFIX +'/simpanintervensi', methods=['POST'])
def simpanintervensi():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        if request.method == 'POST':
            tahun = request.args['tahun']
            tabel = request.args['tabel']
            kode = request.args['kode']    
            indata = dict(request.form)
            sqlcreate = "CREATE TABLE IF NOT EXISTS public.tiv_%s (id BIGSERIAL PRIMARY KEY, kdebps character varying(12), kdepum character varying(12), tahun smallint, intervensi text)" % (tabel) 
            engine.execute(sa_text(sqlcreate).execution_options(autocommit=True))        
            sqldelete = "DELETE FROM tiv_%s where kdebps='%s' and tahun=%s" % (tabel, kode, tahun)
            engine.execute(sa_text(sqldelete).execution_options(autocommit=True))
            htmldata = indata['htmldata'][0]
            isian = "'" + htmldata.replace("'","") + "'"
            print(htmldata)
            sqlinsert = "INSERT INTO tiv_%s (kdebps,tahun,intervensi) values (%s,%s,%s)" % (tabel, kode, tahun, isian)
            print("X",sqlinsert)
            engine.execute(sa_text(sqlinsert))
    return Response("{OK}", mimetype='application/json') 

@app.route(config.PREFIX +'/intervensi', methods=['GET'])
def intervensi():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        tabel = request.args['tabel']
        tahun = request.args['tahun']
        kode = request.args['kode']
        # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
        sqlgethtml = "SELECT * from tiv_%s WHERE tahun=%s AND kdebps='%s'" % (tabel,tahun,kode)
        print(sqlgethtml)
        result = engine.execute(sa_text(sqlgethtml))
        resultkey = engine.execute(sa_text(sqlgethtml)).keys()
        row = result.fetchone()
        print(row)
        output = {}
        try:
            for key, val in zip(resultkey,row):
                output[key] = val
        except:
            pass
        try:
            jsonoutput = json.dumps(output['intervensi'])
        except:
            jsonoutput = json.dumps('')
        # engine.dispose()
        return Response(jsonoutput, mimetype='application/json')

@app.route(config.PREFIX +'/resetpassword', methods=['GET','POST'])
def resetpassword():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        form = LoginForm()
        offset = get_page_args()
        cari = request.args.get('cari')
        cariteks = request.args.get('cariTeks')
        username = request.args['kode']
        if request.method == 'POST':
            user=t_user.query.filter_by(username=username).first()           
            user.password = request.form['password']
            db.session.commit()
            return redirect(url_for('daftar_desa'))  
        else:
            if cari:
                string_cari = '%' + cari + '%'
                rows = t_user.query.filter(or_(t_user.username.ilike(string_cari), t_user.desa.ilike(string_cari), t_user.kecamatan.ilike(string_cari), t_user.kabupaten.ilike(string_cari))).count()
                desa = t_user.query.filter(or_(t_user.username.ilike(string_cari), t_user.desa.ilike(string_cari), t_user.kecamatan.ilike(string_cari), t_user.kabupaten.ilike(string_cari))).order_by(t_user.username.asc()).offset(offset[2]).limit(offset[1])
            else:
                rows = t_user.query.order_by(t_user.username).count()
                desa = t_user.query.order_by(t_user.username.asc()).offset(offset[2]).limit(offset[1])
            per_page = 10
            search = False
            q = request.args.get('q')
            if q:
                search = True
            page = request.args.get('page', type=int, default=1)
            pagination = Pagination(page=page, total=rows, per_page=per_page, search=search, record_name='Desa',css_framework='bootstrap3')
            return render_template('dashboard/resetpassword.htm', landing_logo=landing_logo.config_desc, landing_judul=landing_judul.config_desc,title=landing_judul.config_desc, desa=desa, per_page=per_page,  page=page, pagination=pagination, form=form,username=username)

@app.route(config.PREFIX +'/caridaerah', methods=['GET'])
def caridaerah():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        cari = request.args['cari']
        tipecari = request.args['tipecari']
        if tipecari == 'carikode':
            sqlcari = "SELECT DISTINCT username as kodebps, desa, kecamatan, kabupaten, provinsi FROM t_user WHERE username ILIKE '%%%s%%'" % (cari)
        elif tipecari == 'caridesa':
            sqlcari = "SELECT DISTINCT username as kodebps, desa, kecamatan, kabupaten, provinsi FROM t_user WHERE desa ILIKE '%%%s%%'" % (cari)
        elif tipecari == 'carikab':
            sqlcari = "SELECT DISTINCT username as kodebps, desa, kecamatan, kabupaten, provinsi FROM t_user WHERE kabupaten ILIKE '%%%s%%' AND length(username)=4" % (cari)            
        else:
            sqlcari = "SELECT username as kodebps, desa, kecamatan, kabupaten, provinsi FROM t_user WHERE username ILIKE '%%%s%%' OR desa ILIKE '%%%s%%' OR kecamatan ILIKE '%%%s%%' OR kabupaten ILIKE '%%%s%%' OR provinsi ILIKE '%%%s%%'" % (cari,cari,cari,cari,cari)
        print(sqlcari)
        output = []
        outputline = {}
        result = engine.execute(sa_text(sqlcari))
        for item in result:
            outputline = {}
            outputline['kodebps'] = item[0]
            outputline['desa'] = item[1]
            outputline['kecamatan'] = item[2]
            outputline['kabupaten'] = item[3]
            outputline['provinsi'] = item[4]
            output.append(outputline)
            jsonoutput = json.dumps(output)
            print(item)
        return Response(jsonoutput, mimetype='application/json') 

@app.route(config.PREFIX +'/eksportmplds', methods=['GET'])
def eksportmplds():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        if request.method == 'GET':
            tabel = request.args['tabel']
            sql1 = "SELECT variabel, v_deskripsi FROM tbd_%s" % (tabel)
            result1 = engine.execute(sql1)
            isian = []
            wb = Workbook()
            ws = wb.active
            ws.title = "Keterangan"
            ws1 = wb.create_sheet("Isian",1)
            ws1.title = "Isian"
            ws['A1'] = 'KETERANGAN ISIAN'
            ws['A2'] = '----------------'
            ws['A3'] = 'Berikut adalah keterangan kolom isian.'
            rws = 5
            ws1['A1'] = 'kdebps'
            ws1['B1'] = 'tahun'
            irow = 1
            icol = 3
            for item in result1:
                if item[0] != 'tipe':
                    ws['A' + str(rws)] = item[0]
                    ws['B' + str(rws)] = item[1]
                    rws = rws + 1
                    ws1.cell(row=irow, column=icol, value=item[0])
                    icol = icol+1            
            vwb=save_virtual_workbook(wb)
            output = make_response(vwb)
            rndfile = randomstr_generator()
            wb.save(APP_ROOT + 'app/static/data/tmp/'+ rndfile +'.xlsx')
            return send_from_directory(directory=APP_ROOT + 'app/static/data/tmp/', filename=rndfile +'.xlsx')

@app.route(config.PREFIX +'/impordatads', methods=['POST'])
def impordatads():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        if request.method == 'POST':
            tabel = request.args['tabel']
            if 'qqfile' not in request.files:
                return Response('{"success": false}', mimetype='application/json') 
            file = request.files['qqfile']
            if file.filename == '':
                return Response('{"success": false}', mimetype='application/json') 
            if file:
                filename = secure_filename(file.filename)
                file.save(os.path.join(APP_ROOT + 'app/static/data/tmp/', filename))
                wb = load_workbook(APP_ROOT + 'app/static/data/tmp/' + filename)
                ws = wb['Isian']
                fs = list(ws.rows)[0]
                fsrows = ws.max_row 
                sdrow = 1
                kolom = ''
                isian = ''
                kode = ''
                tahun = ''
                # skor = 0
                for kol in fs:
                    kolom = kolom + kol.value + ', '
                kolom = kolom[:-2]
                # print(kolom)
                while sdrow < fsrows:
                    # print(list(ws.rows)[sdrow])
                    kolval = ''
                    skor = 0
                    koln = 1
                    for kol in list(ws.rows)[sdrow]:
                        # print(kol, kol.value)
                        if koln == 1:
                            kolval = kolval+str(kol.value)+ ', '
                            kode = str(kol.value)
                        if koln == 2:
                            kolval = kolval+ str(kol.value)+ ", "
                            tahun = str(kol.value)     
                        if koln > 2:          
                            kolval = kolval+str(kol.value)+ ', '           
                            # skor = skor + int(kol.value) 
                        koln = koln + 1
                    sdrow = sdrow + 1
                    kolval = kolval[:-2]
                    # print(kolval)
                    try:
                        sqldelete = "DELETE FROM tb_%s where kdebps='%s' and tahun=%s" % (tabel, kode, tahun)
                        # print(sqldelete)
                        engine.execute(sa_text(sqldelete).execution_options(autocommit=True))
                    except:
                        pass
                    sqlinsert = "INSERT INTO tb_%s (%s) values (%s)" % (tabel, kolom, kolval)
                    engine.execute(sa_text(sqlinsert).execution_options(autocommit=True))
                    # print(sqlinsert)
            return Response('{"success": true}', mimetype='application/json') 

@app.route(config.PREFIX +'/isidatadasar')
def isidatadasar():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        # variabel = t_variabel.query.order_by(t_variabel.id.asc()).all()
        tabel = request.args['tabel']
        var1 = "SELECT * from t_datadasar"
        varres = engine.execute(var1)
        variabel = varres.fetchall()
        # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
        provinsi = "select distinct provinsi, left(username, 2) as username from t_user where provinsi != 'ADMIN' and left(username, 2) !='' and username !='0' order by left(username, 2)" 
        kabupaten = "select distinct kabupaten, left(username, 4) as username from t_user where provinsi != 'ADMIN' and left(username, 4) !='' and kabupaten !='-' and username !='0' order by kabupaten"
        desa = "select distinct desa, username from t_user where provinsi != 'ADMIN' order by desa"
        sqldata = "SELECT variabel, v_deskripsi from tbd_%s" % (tabel)
        datan = engine.execute(sqldata)
        datadesc = datan.fetchall()
        l_provinsi = engine.execute(provinsi)
        l_kabupaten = engine.execute(kabupaten)
        l_desa = engine.execute(desa)
        # engine.dispose()
        return render_template('dashboard/isidatadasar.htm', landing_logo=landing_logo.config_desc, landing_judul=landing_judul.config_desc,tabel=tabel,variabel=variabel, provinsi=l_provinsi, kabupaten=l_kabupaten, desa=l_desa, datadesc=datadesc,title=landing_judul.config_desc)

@app.route(config.PREFIX +'/simpandatadasar', methods=['POST'])
def simpandatadasar():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        if request.method == 'POST':
            tabel = request.args['tabel']
            tahun = request.args['tahun']  
            kode = request.args['kode']
            indata = dict(request.form)
            print(tabel,kode,tahun,indata)
            kolom = ''
            isian = ''
            try:
                sqldelete = "DELETE FROM tb_%s where kdebps='%s' and tahun=%s" % (tabel, kode, tahun)
                # print(sqldelete)
                engine.execute(sa_text(sqldelete).execution_options(autocommit=True))
            except:
                pass    
            for items in indata:
                print(items)
                kolom = kolom + items + ', '
                isian = isian + str(indata[items][0]) + ', '
            kolom = kolom[:-2]
            isian = isian[:-2]
            print(kolom, isian)
            sqlinsert = "INSERT INTO tb_%s (kdebps,tahun,%s) values ('%s',%s,%s)" % (tabel,kolom,kode,tahun,isian)
            print(sqlinsert)       
            engine.execute(sa_text(sqlinsert).execution_options(autocommit=True))
        # return Response('{"success": true}', mimetype='application/json') 
        return redirect(url_for('variabel'))

@app.route(config.PREFIX +'/initsetup', methods=['GET','POST'])
def initsetup():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        if request.method == 'POST':
            print(request.files)
            file_kab = request.files['admin_kab_shape']
            file_desa = request.files['admin_desa_shape']
            if file_kab.filename == '':
                flash('Berkas Daerah Administrasi Kabupaten tidak dimasukkan!')
                return redirect(url_for('initsetup')) 
            if file_kab:
                filename_kab = secure_filename(file_kab.filename)
                file_kab.save(os.path.join(APP_ROOT + 'app/static/data/tmp/', filename_kab))
                with zipfile.ZipFile(APP_ROOT + 'app/static/data/tmp/' + filename_kab) as zip_kab:
                    zip_kab.extractall(APP_ROOT + 'app/static/data/tmp/')
            if file_desa.filename == '':
                flash('Berkas Daerah Administrasi Desa tidak dimasukkan!')
                return redirect(url_for('initsetup')) 
            if file_desa:
                filename_desa = secure_filename(file_desa.filename)
                file_desa.save(os.path.join(APP_ROOT + 'app/static/data/tmp/', filename_desa))
                with zipfile.ZipFile(APP_ROOT + 'app/static/data/tmp/' + filename_desa) as zip_kab:
                    zip_kab.extractall(APP_ROOT + 'app/static/data/tmp/')
                admin_desa_dbf = DBF(APP_ROOT + 'app/static/data/tmp/' + filename_desa.split('.')[0] + '.dbf')
                list_kdbbps = []
                tabel = filename_desa.split('.')[0]
                admin_desa_shp = filename_desa.split('.')[0] + '.shp'
                admin_kab_shp = filename_kab.split('.')[0] + '.shp'
                for kdbbps in admin_desa_dbf:
                    list_kdbbps.append(str(int(kdbbps['KDBBPS'])))
                set_kdbbps = set(list_kdbbps)
                os.chdir(APP_ROOT + 'app/static/data/tmp/')
                print(os.getcwd())
                sqldropadmindesa = "DROP TABLE IF EXISTS admin_desa"
                engine.execute(sa_text(sqldropadmindesa).execution_options(autocommit=True))
                sqldropadminkab = "DROP TABLE IF EXISTS admin_kab"
                engine.execute(sa_text(sqldropadminkab).execution_options(autocommit=True))
                processpostgis = 'ogr2ogr -f "PostgreSQL" "PG:host=127.0.0.1 user=%s dbname=%s password=%s" "%s" -lco GEOMETRY_NAME=geom -lco FID=id -lco PRECISION=no -nlt PROMOTE_TO_MULTI -nln admin_desa -overwrite' % (PG_USER, PG_DB, PG_PASSWORD, admin_desa_shp)
                subprocess.call(processpostgis, shell=True)
                processpostgiskab = 'ogr2ogr -f "PostgreSQL" "PG:host=127.0.0.1 user=%s dbname=%s password=%s" "%s" -lco GEOMETRY_NAME=geom -lco FID=id -lco PRECISION=no -nlt PROMOTE_TO_MULTI -nln admin_kab -overwrite' % (PG_USER, PG_DB, PG_PASSWORD, admin_kab_shp)
                subprocess.call(processpostgiskab, shell=True)
                sqltruncpart = "DELETE FROM t_user WHERE username != 'administrator'"
                engine.execute(sa_text(sqltruncpart).execution_options(autocommit=True))
                try:
                    os.remove(APP_ROOT + 'app/static/data/tmp/' +'admin_kab.geojson')
                except:
                    pass
                proceskab = 'ogr2ogr -f "GeoJSON" %s.geojson %s' % ('admin_kab', admin_kab_shp)
                subprocess.call(proceskab, shell=True)
                try:
                    os.remove(APP_ROOT + 'app/static/data/' +'admin_kab.topojson')
                except:
                    pass
                # procestopokab = 'geo2topo %s.geojson > %s.topojson' % ('admin_kab', 'admin_kab')
                procestopokab = 'geo2topo -q 1e5 %s.geojson > ../%s.topojson' % ('admin_kab', 'admin_kab')
                subprocess.call(procestopokab, shell=True)
                for kdbbps in set_kdbbps:
                    print(kdbbps)
                    try:
                        os.remove(APP_ROOT + 'app/static/data/tmp/' + kdbbps + '.geojson')
                    except:
                        pass
                    try:
                        os.remove(APP_ROOT + 'app/static/data/' + kdbbps + '.topojson')
                    except:
                        pass
                    processline = 'ogr2ogr -f "GeoJSON" %s.geojson %s -sql "select * from %s where KDBBPS=\'%s\'"' % (kdbbps,admin_desa_shp,tabel,kdbbps)
                    # print(processline)
                    subprocess.call(processline, shell=True)
                    # procestopo = 'geo2topo %s.geojson > %s.topojson' % (kdbbps, kdbbps)
                    procestopo = 'geo2topo -q 1e5 %s.geojson > ../%s.topojson' % (kdbbps, kdbbps)
                    subprocess.call(procestopo, shell=True)
                    str_find = '"objects":{"%s":' % (kdbbps)
                    str_place = '"objects":{"units":'
                    with open(APP_ROOT + 'app/static/data/' + kdbbps + '.topojson') as f:
                        s = f.read()
                    with open(APP_ROOT + 'app/static/data/' + kdbbps + '.topojson', 'w') as f:
                        s = s.replace(str_find, str_place)
                        f.write(s)
                sqlinsertdesa = "INSERT INTO t_user (username,desa,kecamatan,kabupaten,provinsi,kodepum,password) SELECT kdebps AS username, namobj AS desa, wadmkc AS kecamatan, wadmkk AS kabupaten, wadmpr AS provinsi, REPLACE(kdepum, '.','') AS kodepum, 'rahasia' as password FROM admin_desa"
                engine.execute(sa_text(sqlinsertdesa).execution_options(autocommit=True))
                sqlinsertkab = "INSERT INTO t_user (username,kabupaten,provinsi,kodepum,password) SELECT DISTINCT SUBSTR(kdebps,1,4) AS username, wadmkk AS kabupaten, wadmpr AS provinsi, REPLACE(kdbpum, '.','') AS kodepum, 'rahasia' as password FROM admin_desa WHERE SUBSTR(kdebps,1,4) != '0' GROUP BY SUBSTR(kdebps,1,4), wadmkk, wadmpr, REPLACE(kdbpum, '.','')"
                engine.execute(sa_text(sqlinsertkab).execution_options(autocommit=True))
                sqlinsterprov = "INSERT INTO t_user (username,provinsi,kodepum,password) SELECT DISTINCT SUBSTR(kdebps,1,2) AS username, wadmpr AS provinsi, REPLACE(kdppum, '.','') AS kodepum, 'rahasia' as password FROM admin_desa WHERE SUBSTR(kdebps,1,2) != '0' GROUP BY SUBSTR(kdebps,1,2),wadmpr, REPLACE(kdppum, '.','')"
                engine.execute(sa_text(sqlinsterprov).execution_options(autocommit=True))
                flash('Setup awal selesai, lanjutkan dengan prosedur Landing Page.')
                return redirect(url_for('initsetup')) 
    return render_template('dashboard/initsetup.htm', landing_logo=landing_logo.config_desc, landing_judul=landing_judul.config_desc, title=landing_judul.config_desc)

@app.route(config.PREFIX +'/isipertanian', methods=['POST','GET'])
def isipertanian():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        if request.method == 'POST':
            pass
        else:
            # variabel = t_variabel.query.order_by(t_variabel.id.asc()).all()
            tabel = 'padi'
            var1 = "SELECT * from t_datadasar"
            varres = engine.execute(var1)
            variabel = varres.fetchall()
            # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
            provinsi = "select distinct provinsi, left(username, 2) as username from t_user where provinsi != 'ADMIN' and left(username, 2) !='' and username !='0' order by left(username, 2)" 
            kabupaten = "select distinct kabupaten, left(username, 4) as username from t_user where provinsi != 'ADMIN' and left(username, 4) !='' and kabupaten !='-' and username !='0' order by kabupaten"
            desa = "select distinct desa, username from t_user where provinsi != 'ADMIN' order by desa"
            sqldata = "SELECT variabel, v_deskripsi from tpd_%s" % (tabel)
            datan = engine.execute(sqldata)
            datadesc = datan.fetchall()
            l_provinsi = engine.execute(provinsi)
            l_kabupaten = engine.execute(kabupaten)
            l_desa = engine.execute(desa)
            # engine.dispose()
            return render_template('dashboard/isipertanian.htm', landing_logo=landing_logo.config_desc, landing_judul=landing_judul.config_desc,tabel=tabel,variabel=variabel, provinsi=l_provinsi, kabupaten=l_kabupaten, desa=l_desa, datadesc=datadesc,title=landing_judul.config_desc)

@app.route(config.PREFIX +'/simpanpertanian', methods=['POST'])
def simpanpertanian():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        if request.method == 'POST':
            tabel = request.args['tabel']
            # tahun = request.args['tahun']  
            kode = request.args['kode']
            indata = dict(request.form)
            print(tabel,kode,indata)
            kolom = ''
            isian = ''
            tahun = indata['masa_tanam'][0]
            try:
                sqldelete = "DELETE FROM tp_%s WHERE kdebps='%s' AND masa_tanam='%s'" % (tabel, kode, tahun)
                print(sqldelete)
                engine.execute(sa_text(sqldelete).execution_options(autocommit=True))
            except:
                pass    
            for items in indata:
                print(items)
                kolom = kolom + items + ', '
                isian = isian + str(indata[items][0]) + ', '
            kolom = kolom[:-2]
            isian = isian[:-2]
            print(kolom, isian)
            # sqlinsert = "INSERT INTO tb_%s (kdebps,tahun,%s) values ('%s',%s,%s)" % (tabel,kolom,kode,tahun,isian)
            sqlinsert = "INSERT INTO tp_%s (lahan_tani_ha,tipe_lahan,masa_tanam,luas_tanam_ha,p_urea,p_kcl,p_tsp,p_obat,jenis,kdebps) values (%s,'%s','%s',%s,%s,%s,%s,%s,'%s','%s')" % (tabel,indata['lahan_tani_ha'][0],indata['tipe_lahan'][0],indata['masa_tanam'][0],indata['luas_tanam_ha'][0],indata['p_urea'][0],indata['p_kcl'][0],indata['p_tsp'][0],indata['p_obat'][0],'Padi',kode)
            print(sqlinsert)       
            engine.execute(sa_text(sqlinsert).execution_options(autocommit=True))
            return Response('{"success": true}', mimetype='application/json') 

@app.route(config.PREFIX +'/rekap_padi')
def rekap_padi():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    else:
        offset = get_page_args()
        cari = request.args.get('cari')
        cariteks = request.args.get('cariTeks')
        loggeduser = session.get('username')
        cariuser = '%' + loggeduser + '%'
        print(loggeduser)
        if cari:
            string_cari = '%' + cari + '%'
            # rows = t_user.query.filter(or_(t_user.username.ilike(string_cari), t_user.kodepum.ilike(string_cari), t_user.desa.ilike(string_cari), t_user.kecamatan.ilike(string_cari), t_user.kabupaten.ilike(string_cari))).count()
            # desa = t_user.query.filter(or_(t_user.username.ilike(string_cari), t_user.kodepum.ilike(string_cari), t_user.desa.ilike(string_cari), t_user.kecamatan.ilike(string_cari), t_user.kabupaten.ilike(string_cari))).order_by(t_user.username.asc()).offset(offset[2]).limit(offset[1])            
            desa = engine.execute("SELECT kdebps,jenis,lahan_tani_ha,tipe_lahan,masa_tanam,luas_tanam_ha,p_urea,p_kcl,p_tsp,p_obat FROM tp_padi WHERE kdebps ILIKE '%s' LIMIT %s OFFSET %s" % (string_cari,offset[1],offset[2])).fetchall()
            rows = 0
            for item in desa:
                print(item)
                rows = rows+1
        else:
            desa = engine.execute("SELECT kdebps,jenis,lahan_tani_ha,tipe_lahan,masa_tanam,luas_tanam_ha,p_urea,p_kcl,p_tsp,p_obat FROM tp_padi LIMIT %s OFFSET %s" % (offset[1],offset[2])).fetchall()
            rows = 0
            for item in desa:
                print(item)
                rows = rows+1
            # rows = t_user.query.filter(t_user.username.ilike(cariuser)).count()
            # desa = t_user.query.filter(t_user.username.ilike(cariuser)).offset(offset[2]).limit(offset[1])
        per_page = 10
        search = False
        q = request.args.get('q')
        if q:
            search = True
        page = request.args.get('page', type=int, default=1)
        pagination = Pagination(page=page, total=rows, per_page=per_page, search=search, record_name='Desa',css_framework='bootstrap3')
        return render_template('dashboard/rekap_padi.htm', landing_logo=landing_logo.config_desc, landing_judul=landing_judul.config_desc,title=landing_judul.config_desc, desa=desa, per_page=per_page,  page=page, pagination=pagination)

@app.route(config.PREFIX +'/isian_mobile')
def isian_mobile():
    if not session.get('logged_in'):
        return redirect(url_for('login_mobile'))
    else:    
        variabel = t_variabel.query.order_by(t_variabel.id.asc()).all()
        # engine = create_engine(config.SQLALCHEMY_DATABASE_URI)
        provinsi = "select distinct provinsi, left(username, 2) as username from t_user where provinsi != 'ADMIN' and left(username, 2) !='' and username !='0' order by left(username, 2)" 
        kabupaten = "select distinct kabupaten, left(username, 4) as username from t_user where provinsi != 'ADMIN' and left(username, 4) !='' and kabupaten !='-' and username !='0' order by kabupaten"
        desa = "select distinct desa, username from t_user where provinsi != 'ADMIN' order by desa"
        l_provinsi = engine.execute(provinsi)
        l_kabupaten = engine.execute(kabupaten)
        l_desa = engine.execute(desa)
        # engine.dispose()
        return render_template('dashboard/isivariabel_mobile.htm', landing_logo=landing_logo.config_desc, landing_judul=landing_judul.config_desc,variabel=variabel, provinsi=l_provinsi, kabupaten=l_kabupaten, desa=l_desa,title=landing_judul.config_desc)

@app.route(config.PREFIX +'/login_mobile', methods=['POST','GET'])
def login_mobile():
    form = LoginForm()
    if request.method == 'GET':
        return render_template('dashboard/login_mobile.htm', form=form, landing_logo=landing_logo.config_desc, landing_judul=landing_judul.config_desc,title=landing_judul.config_desc)
    elif request.method == 'POST':
        if form.validate_on_submit():
            user=t_user.query.filter(or_(t_user.username.like(request.form['username']), t_user.kodepum.like(request.form['username']))).first()
            # user=t_user.query.filter_by(username=request.form['username']).first()
            # print(user.username, user.password)
            if user:
                if user.password == request.form['password']:
                    # login_user(user,remember=True)
                    session['logged_in'] = True
                    session['username'] = user.username
                    isnumber = True
                    try:
                        int(user.username)
                    except:
                        isnumber = False
                    if len(user.username) == 2 and isnumber:
                        session['kelas'] = 'provinsi'
                        session['nama'] = user.provinsi
                    if len(user.username) == 4 and isnumber:
                        session['kelas'] = 'kabupaten'
                        session['nama'] = user.kabupaten
                    if len(user.username) == 10 and isnumber:
                        session['kelas'] = 'desa'
                        session['nama'] = user.desa
                    if not isnumber:
                        session['kelas'] = 'ADMIN'
                        session['nama'] = 'ADMIN'
                    flash('Logged')
                    # nextpage = request.args.get('next')
                    # print('NEXT:',nextpage)
                    # if not is_safe_url(next):
                    #     return abort(400)
                    return redirect(url_for('isian_mobile'))       
                else:
                    flash('Password Salah!')      
                    return redirect(url_for('login_mobile'))        
            else:
                flash('User tidak terdaftar')      
                return redirect(url_for('login_mobile'))   
        else:
            flash('Form tidak tervalidasi')
            return redirect(url_for('login_mobile'))  

@app.route(config.PREFIX +"/logout_mobile")
# @login_required
def logout_mobile():
    session['logged_in'] = False
    session['username'] = ''
    session['kelas'] = '' 
    session['nama'] = ''
    # logout_user()
    return redirect(url_for('login_mobile'))

# @app.route(config.PREFIX +'/proxy', methods=['POST','GET'])
# def crossdom():
#     reply = proxypy.get(request.query_string)
#     # print request.query_string
#     # print "REPLY:", reply
#     proxy = json.loads(reply)
#     result = proxy['content']
#     if request.method == 'POST':
#         # print "POST", request.query_string
#         # print "DATA", request.data
#         # print "HEADER", request.headers
#         data = request.data
#         reply = proxypy.post(request.query_string,request.data,request.headers)
#         proxy = json.loads(reply)
#         result = proxy['content']
#     return Response(result,status=200, mimetype='text/plain')

